from __future__ import annotations

import csv
import json
import hashlib
import io
import itertools
import subprocess
import sys
import tempfile
import time
import types
import unittest
from contextlib import redirect_stdout
from datetime import datetime, timedelta, timezone
from pathlib import Path
from unittest import mock


ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "scripts"))

from application_intelligence_agents import (  # noqa: E402
    ApplicationInfoAgent,
    ApplicationIntelligencePipeline,
    FitEvidenceAgent,
    OutreachWriterAgent,
    normalize_publish_time,
    run_pipeline,
)
from codex_runtime_outreach import PROMPT_VERSION  # noqa: E402
from ai_application_workflow import record_needs_completion  # noqa: E402
from run_project_workflow import (  # noqa: E402
    DEFAULT_CANDIDATE_PROFILE,
    PROJECT_ROOT,
    build_workflow_summary,
    can_complete_from_checkpoint,
    checkpoint_body_summary,
    completion_target_ids,
    emit_stage,
    partition_job_ai_targets,
    reuse_completed_records,
    resolve_project_path,
    rewrite_limit,
    rewrite_unlimited_args,
    write_project_manifest,
)
from codex_runtime_outreach import CodexRuntimeOutreachAgent, _prompt  # noqa: E402
import parallel_body_completion as body_completion  # noqa: E402
import run_project_workflow as workflow  # noqa: E402
from parallel_body_completion import (  # noqa: E402
    contains_rate_limit,
    contains_security_verification,
    deduplicate_cards,
    detail_url_candidates,
    record_is_complete,
    record_key,
)

VENDOR_SCRIPTS = ROOT / "vendor" / "xiaohongshu-relay-scrape" / "scripts"
sys.path.insert(0, str(VENDOR_SCRIPTS))
from build_structured_excel import write_csv, write_dedup_workbook  # noqa: E402


TZ = timezone(timedelta(hours=8), name="Asia/Shanghai")
COLLECTED = datetime(2026, 7, 28, 9, 31, 42, tzinfo=TZ)
PROFILE = {
    "name": "示例用户",
    "evidence_items": [
        {
            "id": "github-ai-project",
            "category": "project",
            "label": "AI Agent 项目",
            "detail": "使用 Python、Codex 和 Agent 工作流完成数据分析自动化。",
            "source": "https://github.com/example/project",
        },
        {
            "id": "marketing-experience",
            "category": "experience",
            "label": "内容营销经历",
            "detail": "负责社交媒体内容运营、市场调研和英文沟通。",
            "source": "resume.pdf",
        },
    ],
}


class TimeNormalizationTests(unittest.TestCase):
    def test_relative_and_absolute_time_normalization(self) -> None:
        cases = {
            "昨天 20:05 北京": ("2026-07-27T20:05+08:00", "minute", False),
            "前天 08:10": ("2026-07-26T08:10+08:00", "minute", False),
            "3天前": ("2026-07-25", "day", True),
            "6天前": ("2026-07-22", "day", True),
            "2小时前": ("2026-07-28T07:31+08:00", "relative_hour", True),
            "15分钟前": ("2026-07-28T09:16+08:00", "relative_minute", True),
            "2026-07-27 18:30": ("2026-07-27T18:30+08:00", "minute", False),
            "7月27日 12:00": ("2026-07-27T12:00+08:00", "minute", False),
        }
        for raw, expected in cases.items():
            with self.subTest(raw=raw):
                result = normalize_publish_time(raw, COLLECTED)
                self.assertEqual((result["value"], result["precision"], result["is_estimated"]), expected)

    def test_date_only_does_not_invent_clock_time(self) -> None:
        result = normalize_publish_time("昨天", COLLECTED)
        self.assertEqual(result["value"], "2026-07-27")
        self.assertEqual(result["precision"], "day")
        self.assertTrue(result["is_estimated"])


class CompletionResumeTests(unittest.TestCase):
    def test_checkpoint_completion_can_continue_when_relay_scrape_fails(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            output = Path(temporary)
            (output / "xiaohongshu_cards_latest.json").write_text(
                json.dumps([{"note_id": "complete"}, {"note_id": "missing"}]),
                encoding="utf-8",
            )
            (output / "xiaohongshu_notes_latest.json").write_text(
                json.dumps(
                    [
                        {"note_id": "complete", "body": "full role", "access_status": "detail_ok"},
                        {"note_id": "missing", "body": "", "access_status": "detail_timeout"},
                    ]
                ),
                encoding="utf-8",
            )

            self.assertTrue(can_complete_from_checkpoint(output, True))
            self.assertFalse(can_complete_from_checkpoint(output, False))
            summary = checkpoint_body_summary(output, stop_reason="relay_connection_failed")
            self.assertEqual(summary["bodySucceeded"], 1)
            self.assertEqual(summary["missingBodies"], 1)
            self.assertTrue(summary["checkpointFallback"])

    def test_reuses_completed_outputs_but_enriches_newly_completed_body(self) -> None:
        previous = {
            "records": [
                {
                    "note_id": "new-body",
                    "body": "",
                    "job_card": {"parse_basis": "search_card"},
                    "outreach": {"runtime_status": "fallback_missing_job_body"},
                },
                {
                    "note_id": "already-complete",
                    "body": "完整岗位正文",
                    "job_card": {"parse_basis": "full_body", "marker": "preserved"},
                    "outreach": {"runtime_status": "completed"},
                },
            ],
        }
        payload = {
            "records": [
                {
                    "note_id": "new-body",
                    "body": "本次续跑新抓到的完整岗位正文",
                    "job_card": {"parse_basis": "full_body"},
                    "outreach": {"runtime_status": "completed"},
                },
                {
                    "note_id": "already-complete",
                    "body": "完整岗位正文",
                    "job_card": {"parse_basis": "full_body"},
                    "outreach": {"runtime_status": "completed"},
                },
            ],
        }

        reused = reuse_completed_records(payload, previous)
        targets = completion_target_ids(payload, previous)

        self.assertEqual(reused, 1)
        self.assertEqual(payload["records"][1]["job_card"]["marker"], "preserved")
        self.assertNotIn("marker", payload["records"][0]["job_card"])
        self.assertEqual(targets, {"new-body"})

    def test_reuses_cached_image_text_while_retrying_incomplete_record(self) -> None:
        previous = {
            "records": [{
                "note_id": "poster-only",
                "body": "",
                "media": {
                    "analysis": {
                        "status": "analyzed",
                        "source": "vision_model",
                        "visible_text": "项目候选人要求\n熟悉 SQL 和 Excel",
                    },
                },
                "job_card": {"parse_basis": "search_card"},
                "outreach": {"runtime_status": "fallback_missing_job_body"},
            }],
        }
        payload = {
            "records": [{
                "note_id": "poster-only",
                "body": "",
                "media": {"images": [{"url": "https://example.test/poster.webp"}]},
                "job_card": {"parse_basis": "search_card"},
                "outreach": {"runtime_status": "fallback_missing_job_body"},
            }],
        }

        reused = reuse_completed_records(payload, previous)

        self.assertEqual(reused, 0)
        self.assertEqual(
            payload["records"][0]["media"]["analysis"]["visible_text"],
            "项目候选人要求\n熟悉 SQL 和 Excel",
        )


    def test_resume_helpers_match_url_aliases_by_canonical_note_id(self) -> None:
        previous = {
            "records": [{
                "note_url": "https://example.test/search_result/stable-note?token=old",
                "body": "complete role body",
                "job_card": {"parse_basis": "full_body", "marker": "preserved"},
                "outreach": {"runtime_status": "completed"},
            }],
        }
        payload = {
            "records": [{
                "note_url": "https://example.test/explore/stable-note?token=new",
                "body": "complete role body",
                "job_card": {"parse_basis": "full_body"},
                "outreach": {"runtime_status": "completed"},
            }],
        }

        reused = reuse_completed_records(payload, previous)
        targets = completion_target_ids(payload, previous)

        self.assertEqual(reused, 1)
        self.assertEqual(targets, set())
        self.assertEqual(payload["records"][0]["job_card"]["marker"], "preserved")

    def test_candidate_evidence_gap_does_not_mark_job_information_incomplete(self) -> None:
        record = {
            "note_id": "complete-job-card",
            "body": "complete job body",
            "job_card": {"parse_basis": "full_body"},
            "outreach": {"runtime_status": "fallback_missing_candidate_evidence"},
        }

        self.assertFalse(record_needs_completion(record))


class ApplicationAgentTests(unittest.TestCase):
    def test_pipeline_matches_url_only_card_and_note_aliases(self) -> None:
        cards = [{
            "search_result_url": "https://example.test/search_result/stable-note?token=card",
            "title": "AI product intern",
        }]
        notes = [{
            "note_url": "https://example.test/explore/stable-note?token=detail",
            "title": "AI product intern",
            "body": "Own product research, requirements, and delivery.",
            "access_status": "detail_ok",
            "scraped_at": "2026-07-28T09:31:42+08:00",
        }]

        result = ApplicationIntelligencePipeline(PROFILE, now=COLLECTED).run(cards, notes)

        self.assertEqual(len(result.payload["records"]), 1)
        self.assertEqual(result.payload["records"][0]["note_id"], "stable-note")
        self.assertEqual(result.payload["records"][0]["body"], notes[0]["body"])

    def test_outreach_never_uses_card_author_or_publish_time_as_salutation(self) -> None:
        writer = OutreachWriterAgent({
            "candidate_application": {
                "name": "示例候选人",
                "school": "示例大学",
                "major": "公共管理",
                "availabilityDays": "5",
                "internshipDuration": "6个月",
            },
        })
        greeting = writer.run(
            {
                "title": "招聘医学背景实习生",
                "card_author": "小周同学 48分钟前",
                "author": "小周同学",
                "job_card": {"role_name": "实习医学信息沟通员"},
            },
            {
                "responsibilities": [{"text": "负责市场数据收集和整理"}],
                "requirements": [{"text": "沟通能力和执行力强"}],
            },
            [{
                "id": "project-1",
                "label": "市场调研项目",
                "detail": "负责市场数据收集和整理；协同团队完成结果汇报",
                "category": "project",
            }],
        )["greeting"]

        self.assertTrue(greeting.startswith("您好，我是示例候选人"))
        self.assertIn("实习医学信息沟通员", greeting)
        self.assertIn("是否仍在招聘", greeting)
        self.assertNotIn("小周同学", greeting)
        self.assertNotIn("48分钟前", greeting)

    def test_outreach_uses_fixed_format_and_never_turns_a_skill_into_experience(self) -> None:
        writer = OutreachWriterAgent({
            "candidate_application": {
                "name": "示例候选人",
                "school": "示例大学",
                "major": "商业分析",
                "degreeYear": "研二",
                "availabilityDays": "5",
                "internshipDuration": "6个月",
            },
        })

        result = writer.run(
            {"job_card": {"role_name": "商业分析实习生"}},
            {"responsibilities": [], "requirements": []},
            [{"id": "skills-1", "category": "skills", "label": "skills", "detail": "R"}],
        )

        self.assertEqual(result["format_version"], "fixed-cn-application-v2")
        self.assertEqual(result["used_evidence_ids"], [])
        self.assertEqual(result["email_subject"], "应聘商业分析实习生｜示例候选人｜每周可实习5天")
        self.assertTrue(result["greeting"].startswith("您好，我是示例候选人"))
        self.assertTrue(result["email_body"].startswith("尊敬的招聘负责人：\n您好！我是示例候选人"))
        self.assertIn("简历随信附上", result["email_body"])
        self.assertIn("此致\n敬礼！", result["cover_letter"])
        self.assertNotIn("skills相关实践", str(result))
        self.assertNotIn("我R", str(result))

    def test_outreach_uses_verified_first_person_claim_verbatim(self) -> None:
        writer = OutreachWriterAgent({
            "candidate_application": {
                "name": "示例候选人",
                "school": "示例大学",
                "major": "商业分析",
                "degreeYear": "研二",
                "availabilityDays": "5",
                "internshipDuration": "6个月",
            },
        })
        claim = "我负责整理业务数据、分析关键指标并输出周报"
        result = writer.run(
            {"job_card": {"role_name": "数据分析实习生"}},
            {"responsibilities": [{"text": "整理业务数据并分析关键指标"}], "requirements": []},
            [{
                "id": "project-verified",
                "category": "project",
                "label": "业务分析项目",
                "detail": "负责整理业务数据、分析关键指标并输出周报",
                "first_person_claim": claim,
                "source": "resume.txt",
                "source_evidence": "负责整理业务数据、分析关键指标并输出周报",
                "confidence": 96,
            }],
        )

        self.assertIn(f"{claim}。", result["email_body"])
        self.assertIn(f"{claim}。", result["cover_letter"])
        self.assertEqual(result["used_evidence_ids"], ["project-verified"])
        self.assertNotIn("resume.txt", str(result))
        self.assertNotIn("我我", str(result))

    def test_business_analysis_matching_prefers_complete_experience_over_skill_tag(self) -> None:
        agent = FitEvidenceAgent({
            "experiences": [{
                "id": "exp-research",
                "title": "市场研究实习生",
                "actions": ["开展用户需求调研、数据分析并完成竞品分析"],
                "results": ["输出市场研究报告和渠道建议"],
            }],
            "projects": [{
                "id": "project-general-market",
                "name": "市场合作项目",
                "actions": ["推进合作方沟通并整理项目资料"],
                "results": ["完成项目交付"],
            }],
            "skills": ["R", "数据分析"],
        })

        evidence = agent.run(
            {"title": "深圳/香港实习", "job_card": {"role_name": "商业分析实习生（咨询业务）"}},
            [],
        )

        self.assertTrue(evidence)
        self.assertEqual(evidence[0]["id"], "exp-research")
        self.assertNotEqual(evidence[0]["category"], "skills")

        outreach = OutreachWriterAgent({
            "candidate_application": {
                "name": "示例候选人",
                "school": "示例大学",
                "major": "商业分析",
                "degreeYear": "研二",
                "availabilityDays": "5",
                "internshipDuration": "6个月",
            },
        }).run(
            {"job_card": {"role_name": "商业分析实习生（咨询业务）"}},
            {"responsibilities": [], "requirements": []},
            evidence,
        )

        self.assertIn("岗位所需的信息分析与结论输出直接相关", outreach["greeting"])
        self.assertIn("我会先确认业务问题和判断口径", outreach["cover_letter"])
        self.assertEqual(outreach["used_evidence_ids"], ["exp-research"])
        self.assertNotIn("岗位关注的项目", str(outreach))
        self.assertNotIn("我会先对齐目标和交付标准", str(outreach))

    def test_evidence_summary_prefers_role_relevant_action_without_repeating_context(self) -> None:
        evidence = {
            "label": "市场营销实习生",
            "detail": (
                "撰写联动Brief，为活动策划提供决策基础；"
                "围绕竞品分析、KOL追踪和玩家反馈开展数据抓取；"
                "围绕竞品分析、KOL追踪和玩家反馈完成50+次爬取"
            ),
        }
        summary = OutreachWriterAgent._evidence_sentence(
            evidence,
            "数据分析实习：开展用户反馈分析并输出可落地报告",
        )

        self.assertIn("开展数据抓取；完成50+次爬取", summary)
        self.assertEqual(summary.count("围绕竞品分析、KOL追踪和玩家反馈"), 1)
        self.assertEqual(OutreachWriterAgent._evidence_lead(evidence), "我")
        self.assertNotIn("在市场营销实习期间", f"{OutreachWriterAgent._evidence_lead(evidence)}{summary}")

    def test_extracts_application_data_with_provenance(self) -> None:
        note = {
            "body": "岗位职责：负责内容运营和数据分析。要求英语好，熟练使用Codex，每周到岗4天。简历投递：jobs@example.com，也可以私信我。微信：hire_2026",
            "title": "AI产品运营实习",
        }
        result = ApplicationInfoAgent().run(note)
        self.assertEqual({item["type"] for item in result["contacts"]}, {"email", "wechat"})
        self.assertTrue(result["requirements"])
        self.assertTrue(result["responsibilities"])
        self.assertTrue(all(item["source_field"] and item["evidence"] for item in result["contacts"]))
        self.assertTrue(all(item["source_field"] == "body" for item in result["requirements"]))
        requirement_text = {item["text"] for item in result["requirements"]}
        responsibility_text = {item["text"] for item in result["responsibilities"]}
        self.assertFalse(requirement_text & responsibility_text)
        self.assertTrue(all(0.5 <= item["classification_confidence"] <= 1 for item in result["requirements"] + result["responsibilities"]))

    def test_numbered_sections_without_sentence_punctuation_keep_role_boundaries(self) -> None:
        body = (
            "AI产品经理实习 职位描述 "
            "1、负责智能平台的产品设计与优化，产出需求文档和原型图 "
            "2、与研发、策略、算法团队紧密合作，共同推动方案落地 "
            "3、关注用户反馈、业务诉求及运营数据 "
            "4、持续跟踪行业前沿发展并挖掘新场景 "
            "任职要求 "
            "1、本科及以上学历，计算机、统计学相关专业优先 "
            "2、具备良好的沟通能力、抗压能力和责任心 "
            "3、善于学习思考，逻辑思维严密"
        )

        result = ApplicationInfoAgent().run({"body": body})

        self.assertEqual(len(result["responsibilities"]), 4)
        self.assertEqual(len(result["requirements"]), 3)
        self.assertTrue(any(item["text"].startswith("与研发") for item in result["responsibilities"]))
        self.assertTrue(any(item["text"].startswith("本科及以上") for item in result["requirements"]))
        for item in result["responsibilities"] + result["requirements"]:
            self.assertEqual(body[item["offset_start"] : item["offset_end"]], item["text"])

    def test_inline_heading_phrase_does_not_reclassify_following_responsibilities(self) -> None:
        body = (
            "岗位职责："
            "1、负责梳理岗位要求：并同步职位信息 "
            "2、跟进候选人沟通并支持招聘流程"
        )

        result = ApplicationInfoAgent().run({"body": body})

        self.assertEqual(
            [item["text"] for item in result["responsibilities"]],
            ["负责梳理岗位要求：并同步职位信息", "跟进候选人沟通并支持招聘流程"],
        )
        self.assertEqual(result["requirements"], [])
        for item in result["responsibilities"]:
            self.assertEqual(body[item["offset_start"] : item["offset_end"]], item["text"])

    def test_structural_heading_connectors_preserve_section_boundaries(self) -> None:
        body = (
            "岗位要求如下："
            "1、熟悉 Python "
            "2、拥有 3.5 年相关经验 "
            "工作职责包括："
            "1、负责需求分析 "
            "2、跟进项目上线"
        )

        result = ApplicationInfoAgent().run({"body": body})

        self.assertEqual(
            [item["text"] for item in result["requirements"]],
            ["熟悉 Python", "拥有 3.5 年相关经验"],
        )
        self.assertEqual(
            [item["text"] for item in result["responsibilities"]],
            ["负责需求分析", "跟进项目上线"],
        )
        extracted = result["requirements"] + result["responsibilities"]
        self.assertFalse(any(item["text"] in {"如下", "包括"} for item in extracted))
        for item in extracted:
            self.assertEqual(body[item["offset_start"] : item["offset_end"]], item["text"])

    def test_decimal_experience_is_not_treated_as_a_numbered_list_marker(self) -> None:
        body = "任职要求：拥有 3.5 年相关经验。"

        result = ApplicationInfoAgent().run({"body": body})

        self.assertEqual(
            [item["text"] for item in result["requirements"]],
            ["拥有 3.5 年相关经验。"],
        )
        requirement = result["requirements"][0]
        self.assertEqual(body[requirement["offset_start"] : requirement["offset_end"]], requirement["text"])

    def test_complete_pipeline_uses_only_loaded_evidence(self) -> None:
        card = {"note_id": "n1", "title": "AI产品运营实习", "note_url": "https://example/n1"}
        note = {
            **card,
            "body": "要求英语好，擅长Codex和AI工具。请投递到 jobs@example.com",
            "publish_time": "昨天 20:05 北京",
            "scraped_at": "2026-07-28T09:31:42",
            "access_status": "detail_ok",
        }
        result = ApplicationIntelligencePipeline(PROFILE, now=COLLECTED).run([card], [note])
        self.assertTrue(result.passed)
        record = result.payload["records"][0]
        self.assertEqual(record["publish_time"]["value"], "2026-07-27T20:05+08:00")
        self.assertIn("Codex", record["outreach"]["email_body"])
        self.assertTrue(set(record["outreach"]["used_evidence_ids"]) <= {"github-ai-project", "marketing-experience"})

    def test_unmatched_note_does_not_claim_general_background_as_role_evidence(self) -> None:
        profile = {
            "name": "示例用户",
            "evidence_items": [
                {
                    "id": "resume-skills",
                    "category": "skills",
                    "label": "英语、数据与内容工具",
                    "detail": "三份简历共同确认的英语、Excel 和内容工具背景。",
                    "source": "resume.pdf",
                }
            ],
        }
        card = {"note_id": "n1", "title": "法务实习", "note_url": "https://example/n1"}
        note = {
            **card,
            "body": "请按正文中的方式申请。",
            "scraped_at": "2026-07-28T09:31:42",
            "access_status": "detail_ok",
        }
        result = ApplicationIntelligencePipeline(profile, now=COLLECTED).run([card], [note])
        self.assertFalse(result.passed)
        record = result.payload["records"][0]
        outreach = record["outreach"]
        self.assertEqual(record["fit_evidence"], [])
        self.assertEqual(outreach["status"], "needs_review")
        self.assertIn("简历随信附上", outreach["email_body"])
        self.assertNotIn("附件", outreach["email_body"])
        self.assertNotIn("岗位方向与我的经历", outreach["greeting"])

    def test_role_match_requires_specific_terms_and_reports_confidence(self) -> None:
        agent = FitEvidenceAgent(PROFILE)
        matched = agent.run(
            {"title": "AI产品运营实习", "body": "负责Codex工作流和数据分析"},
            [{"text": "熟悉Python"}],
        )
        self.assertTrue(matched)
        self.assertTrue(all(item["match_confidence"] >= 0.68 for item in matched))
        self.assertTrue(all(item["matched_terms"] for item in matched))

        unrelated = agent.run({"title": "法务实习", "body": "协助合同归档"}, [])
        self.assertEqual(unrelated, [])

    def test_user_insight_role_prioritizes_interview_and_survey_evidence(self) -> None:
        agent = FitEvidenceAgent({
            "evidence_items": [
                {
                    "id": "user-research-520",
                    "category": "project",
                    "label": "用户需求调研",
                    "detail": "通过1v1访谈及问卷系统收集520位用户反馈，归纳5类用户需求。",
                    "source": "resume.pdf",
                },
                {
                    "id": "community-150",
                    "category": "project",
                    "label": "社群运营",
                    "detail": "从0到1搭建并运营150人社群，策划内容与活动。",
                    "source": "resume.pdf",
                },
            ],
        })

        matched = agent.run(
            {
                "title": "AI产品运营实习",
                "body": "分析用户query、用户痛点与高频场景，沉淀案例库。",
            },
            [],
        )

        self.assertTrue(matched)
        self.assertEqual(matched[0]["id"], "user-research-520")
        self.assertTrue({"用户需求", "用户反馈", "访谈", "问卷"} & set(matched[0]["matched_terms"]))

    def test_ai_product_role_selects_product_insight_and_operations_axes(self) -> None:
        agent = FitEvidenceAgent({
            "evidence_items": [
                {
                    "id": "asteria-product",
                    "category": "project",
                    "label": "Asteria 数据分析交付系统",
                    "detail": "从0到1搭建Asteria数据分析交付系统，把分散分析纳入统一产品链路，覆盖7类输入和6类场景。",
                    "source": "resume.pdf",
                },
                {
                    "id": "research-applied",
                    "category": "experience",
                    "label": "用户洞察转化",
                    "detail": "开展1v1深访并沉淀4类用户需求与场景，将用户洞察转化为内容策略与栏目结构。",
                    "source": "resume.pdf",
                },
                {
                    "id": "research-volume",
                    "category": "experience",
                    "label": "用户需求调研",
                    "detail": "通过访谈和问卷收集520位用户反馈，归纳5类用户需求。",
                    "source": "resume.pdf",
                },
                {
                    "id": "community-ops",
                    "category": "experience",
                    "label": "社群运营",
                    "detail": "从0到1搭建并运营150人社群，通过话题、内容与活动形成固定运营机制。",
                    "source": "resume.pdf",
                },
                {
                    "id": "agent-backend",
                    "category": "project",
                    "label": "Agent 后端模块",
                    "detail": "使用 Python 开发大模型 Agent 后端模块并完成接口联调。",
                    "source": "resume.pdf",
                },
            ],
        })

        matched = agent.run(
            {
                "title": "AI产品运营实习",
                "body": "负责数据分析产品 BA Agent，分析用户 query 和高频场景，推进社群活动并监控指标。",
            },
            [],
        )

        self.assertEqual(
            [(item["id"], item["role_axis"]) for item in matched],
            [
                ("asteria-product", "ai_product"),
                ("research-applied", "user_insight"),
                ("community-ops", "operations"),
            ],
        )

        engineering = agent.run(
            {"title": "Agent 研发实习", "body": "使用 Python 开发大模型 Agent 后端模块。"},
            [],
        )
        self.assertTrue(engineering)
        self.assertTrue(all(not item.get("role_axis") for item in engineering))

    def test_card_fallback_is_not_counted_as_full_body(self) -> None:
        card = {"note_id": "n1", "title": "AI实习", "note_url": "https://example/n1"}
        note = {
            **card,
            "body": "搜索卡片摘要",
            "scraped_at": "2026-07-28T09:31:42",
            "access_status": "detail_timeout",
        }
        result = ApplicationIntelligencePipeline(PROFILE, now=COLLECTED).run([card], [note])
        self.assertFalse(result.passed)
        self.assertEqual(result.payload["quality_gate"]["body_count"], 0)
        self.assertEqual(result.payload["records"], [])
        self.assertEqual(result.payload["publication_contract"]["published_count"], 0)
        self.assertEqual(result.payload["publication_contract"]["pending_body_count"], 1)
        self.assertFalse(result.payload["quality_gate"]["checks"]["all_discovered_notes_have_records"])
        self.assertTrue(result.payload["quality_gate"]["checks"]["all_scraped_jobs_have_job_cards"])
        self.assertTrue(result.payload["quality_gate"]["checks"]["all_scraped_jobs_have_application_copy"])

    def test_codex_runtime_applies_structured_per_link_output(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            output = Path(temporary)
            fake_cli = output / "codex.cmd"
            fake_cli.write_text("@echo off\n", encoding="utf-8")

            def fake_run(command, **kwargs):
                response_path = Path(command[command.index("--output-last-message") + 1])
                response_path.write_text(
                    json.dumps({"items": [{
                        "note_id": "n1",
                        "greeting": "您好，我是示例用户，想应聘内容运营实习。我曾负责社交媒体内容运营、市场调研和英文沟通。请问岗位目前是否仍在招聘？",
                        "email_subject": "应聘内容运营实习｜示例用户",
                        "email_body": "您好，我希望申请内容运营实习。我在内容营销经历中负责社交媒体内容运营、市场调研和英文沟通，能够对应岗位对内容策划与数据分析的要求。期待进一步沟通岗位当前最需要推进的任务。",
                        "cover_letter": "主题：应聘内容运营实习｜示例用户\n尊敬的招聘负责人：\n您好！我是示例用户，希望申请内容运营实习。我在内容营销经历中负责社交媒体内容运营、市场调研和英文沟通，能够把信息整理、内容判断和协作沟通连接起来，支持岗位所需的内容策划与数据分析。在具体工作中，我会先根据目标受众梳理选题方向，再结合调研反馈判断内容重点，并与相关成员确认发布节奏和交付标准。\n\n针对该岗位，我会先理解团队当前内容目标和数据口径，再从一个具体选题或活动开始验证判断：整理用户反馈和内容表现，识别需要优先优化的环节，把分析结论转化为可执行的内容动作，并在发布后继续复盘结果。这套方法来自我已有的内容营销实践，不依赖未经验证的工具或成果。\n\n我希望进一步了解团队当前最需要推进的内容任务，并具体沟通我可以优先承担的选题研究、内容执行或数据复盘工作。感谢您的阅读，期待进一步沟通。\n\n此致\n敬礼！\n姓名：示例用户",
                        "used_evidence_ids": ["marketing-experience"],
                        "requirement_matches": ["内容策划与数据分析 -> marketing-experience -> 社交媒体内容运营与市场调研"],
                        "recommended_resume": "用户运营",
                        "resume_reason": "岗位核心职责为内容运营。",
                    }]}, ensure_ascii=False),
                    encoding="utf-8",
                )
                return subprocess.CompletedProcess(command, 0, "", "")

            record = {
                "note_id": "n1",
                "note_url": "https://example/n1",
                "title": "内容运营实习",
                "body": "负责内容策划，要求具备数据分析能力。",
                "publish_time": {"value": "2026-07-22"},
                "application_info": {"responsibilities": [], "requirements": [], "application_routes": [], "contacts": []},
                "fit_evidence": [PROFILE["evidence_items"][1]],
                "quality": {"body_present": True},
                "outreach": {},
            }
            agent = CodexRuntimeOutreachAgent(output, candidate_name="示例用户", cli_bin=str(fake_cli), run_command=fake_run)
            report = agent.enrich([record])
            self.assertEqual(report.generated, 1)
            self.assertEqual(record["outreach"]["generation_mode"], "codex_cli_runtime")
            self.assertEqual(record["outreach"]["recommended_resume"], "用户运营")
            self.assertEqual(record["outreach"]["resume_reason"], "岗位核心职责为内容运营。")

    def test_codex_runtime_rejects_meta_or_unstructured_copy(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            output = Path(temporary)
            fake_cli = output / "codex.cmd"
            fake_cli.write_text("@echo off\n", encoding="utf-8")
            agent = CodexRuntimeOutreachAgent(output, candidate_name="示例用户", cli_bin=str(fake_cli), run_command=lambda *args, **kwargs: None)
            source = {
                "note_id": "n1",
                "candidate_evidence": [{"id": "marketing-experience"}],
            }
            bad = {
                "note_id": "n1",
                "greeting": "您好，我想申请这个岗位，附件是我的简历。",
                "email_subject": "申请岗位",
                "email_body": "您好，附件是我的简历，请查收。",
                "cover_letter": "您好，附件是我的简历，请查收。",
                "used_evidence_ids": ["marketing-experience"],
                "requirement_matches": [],
            }
            with self.assertRaises(ValueError):
                agent._validate_output(bad, source)

    def test_cover_letter_prompt_includes_runtime_candidate_profile(self) -> None:
        prompt = _prompt(
            [{"note_id": "n1", "title": "Data analyst intern", "body": "Role evidence."}],
            "",
            {
                "name": "Example Candidate",
                "school": "Example University",
                "major": "Data Analytics",
                "degreeYear": "Year 2",
                "phoneWeChat": "contact-placeholder",
                "email": "candidate@example.com",
                "availabilityDays": "5",
                "internshipDuration": "6 months",
            },
        )
        self.assertIn("Example Candidate", prompt)
        self.assertIn("Example University", prompt)
        self.assertIn("candidate@example.com", prompt)
        self.assertIn("每周可实习", prompt)
        self.assertIn("6 months", prompt)

    def test_quality_gate_detects_partial_collection_and_missing_body(self) -> None:
        cards = [
            {"note_id": "n1", "title": "one"},
            {"note_id": "n2", "title": "two"},
        ]
        notes = [
            {
                "note_id": "n1",
                "title": "one",
                "body": "",
                "scraped_at": "2026-07-28T09:31:42",
            }
        ]
        result = ApplicationIntelligencePipeline(PROFILE, now=COLLECTED).run(cards, notes)
        self.assertFalse(result.passed)
        gate = result.payload["quality_gate"]
        self.assertEqual(gate["discovered_count"], 2)
        self.assertEqual(gate["covered_discovered_count"], 0)
        self.assertFalse(gate["checks"]["all_discovered_notes_have_records"])
        self.assertTrue(gate["checks"]["all_records_have_bodies"])
        self.assertEqual(result.payload["records"], [])
        self.assertEqual(gate["job_cards_generated"], 0)
        self.assertEqual(gate["application_copy_generated"], 0)
        self.assertEqual(result.payload["publication_contract"]["pending_body_count"], 2)
        self.assertTrue(gate["checks"]["all_scraped_jobs_have_job_cards"])
        self.assertTrue(gate["checks"]["all_scraped_jobs_have_application_copy"])

    def test_pipeline_ignores_stale_notes_outside_current_card_checkpoint(self) -> None:
        cards = [
            {"note_id": "current-1", "title": "岗位一"},
            {"note_id": "current-2", "title": "岗位二"},
        ]
        notes = [
            {
                "note_id": "current-1",
                "title": "岗位一",
                "body": "负责数据分析。",
                "scraped_at": "2026-07-28T09:31:42",
                "access_status": "detail_ok",
            },
            {
                "note_id": "stale-1",
                "title": "旧检查点岗位",
                "body": "不应进入当前结果。",
                "scraped_at": "2026-07-28T09:31:42",
                "access_status": "detail_ok",
            },
        ]

        result = ApplicationIntelligencePipeline(PROFILE, now=COLLECTED).run(cards, notes)

        self.assertEqual([record["note_id"] for record in result.payload["records"]], ["current-1"])
        self.assertEqual(result.payload["quality_gate"]["discovered_count"], 2)
        self.assertEqual(result.payload["quality_gate"]["record_count"], 1)
        self.assertEqual(result.payload["quality_gate"]["covered_discovered_count"], 1)
        self.assertEqual(result.payload["publication_contract"]["pending_body_count"], 1)

    def test_csv_export_accepts_image_fields_added_after_first_row(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            output = Path(temporary) / "notes.csv"
            workbook_output = Path(temporary) / "notes.xlsx"
            write_csv(
                output,
                [
                    {"note_id": "n1", "title": "文本岗位", "body": "正文\x05含控制字符"},
                    {
                        "note_id": "n2",
                        "title": "图片岗位",
                        "detail_image_urls": ["https://example.test/job.webp"],
                        "detail_image_alts": ["岗位详情"],
                    },
                ],
            )
            with output.open("r", encoding="utf-8-sig", newline="") as handle:
                rows = list(csv.DictReader(handle))

            self.assertIn("detail_image_urls", rows[0])
            self.assertIn("detail_image_alts", rows[0])
            self.assertIn("job.webp", rows[1]["detail_image_urls"])

            records = [
                {"note_id": "n1", "title": "文本岗位", "body": "正文\x05含控制字符"},
                {
                    "note_id": "n2",
                    "title": "图片岗位",
                    "detail_image_urls": ["https://example.test/job.webp"],
                    "detail_image_alts": ["岗位详情"],
                },
            ]
            write_dedup_workbook(records, workbook_output, output)
            from openpyxl import load_workbook

            workbook = load_workbook(workbook_output, read_only=True)
            raw = workbook["Raw"]
            headers = [cell.value for cell in raw[1]]
            self.assertIn("detail_image_urls", headers)
            self.assertNotIn("\x05", raw.cell(row=2, column=headers.index("body") + 1).value)
            self.assertIn("job.webp", raw.cell(row=3, column=headers.index("detail_image_urls") + 1).value)
            workbook.close()

    def test_writes_all_artifacts(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            output = Path(temporary)
            (output / "xiaohongshu_cards_latest.json").write_text(
                json.dumps([{"note_id": "n1", "title": "AI实习"}], ensure_ascii=False), encoding="utf-8"
            )
            (output / "xiaohongshu_notes_latest.json").write_text(
                json.dumps(
                    [
                        {
                            "note_id": "n1",
                            "title": "AI实习",
                            "body": "\x05=需要Python经验，邮箱 jobs@example.com",
                            "scraped_at": "2026-07-28T09:31:42",
                        }
                    ],
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )
            profile_path = output / "candidate.json"
            profile_path.write_text(json.dumps(PROFILE, ensure_ascii=False), encoding="utf-8")
            result = run_pipeline(output, profile_path, now=COLLECTED)
            self.assertTrue(result.passed)
            self.assertEqual(result.payload["codex_runtime"]["prompt_version"], PROMPT_VERSION)
            for name in (
                "application_intelligence.json",
                "application_intelligence.csv",
                "application_intelligence.xlsx",
                "application_intelligence_summary.json",
                "application_intelligence_report.md",
                "coverage_report.json",
            ):
                self.assertTrue((output / name).is_file(), name)
            payload = json.loads((output / "application_intelligence.json").read_text(encoding="utf-8"))
            self.assertTrue(payload["records"][0]["body"].startswith("\x05="))
            from openpyxl import load_workbook

            workbook = load_workbook(output / "application_intelligence.xlsx", read_only=True)
            self.assertEqual(workbook["Applications"]["J2"].value[:2], "'=")
            workbook.close()


class WorkflowWrapperTests(unittest.TestCase):
    def test_limit_is_always_rewritten_to_unlimited(self) -> None:
        self.assertEqual(
            rewrite_unlimited_args(["--limit", "20", "--max-age-days", "30", "--max-scrolls", "40"]),
            ["--max-age-days", "30", "--max-scrolls", "40", "--limit", "0"],
        )
        self.assertEqual(
            rewrite_unlimited_args(["--limit=200", "--max-age-days=90"]),
            ["--max-age-days=90", "--limit", "0"],
        )

    def test_bootstrap_limit_is_internal_and_bounded(self) -> None:
        self.assertEqual(
            rewrite_limit(["--limit", "20", "--max-scrolls", "40"], 1),
            ["--max-scrolls", "40", "--limit", "1"],
        )

    def test_parallel_completion_requires_full_detail_body(self) -> None:
        complete = {"note_id": "n1", "body": "full body", "access_status": "detail_ok"}
        fallback = {"note_id": "n1", "body": "card text", "access_status": "detail_timeout"}
        legacy_false_success = {
            "note_id": "n1",
            "body": "访问频繁，请稍后再试",
            "access_status": "detail_ok",
        }
        self.assertEqual(record_key(complete), "n1")
        self.assertTrue(record_is_complete(complete))
        self.assertFalse(record_is_complete(fallback))
        self.assertFalse(record_is_complete(legacy_false_success))
        self.assertTrue(contains_security_verification("请完成安全验证后继续"))
        self.assertTrue(contains_rate_limit("访问频繁，请稍后再试 300013"))
        self.assertFalse(contains_security_verification("访问频繁，请稍后再试"))
        self.assertFalse(contains_security_verification("岗位详情正常展示"))

    def test_parallel_completion_prioritizes_and_deduplicates_detail_urls(self) -> None:
        card = {
            "explore_url": "https://example.com/explore/n1",
            "search_result_url": "https://example.com/search/n1",
            "note_url": "https://example.com/explore/n1",
        }
        self.assertEqual(
            detail_url_candidates(card),
            ["https://example.com/search/n1", "https://example.com/explore/n1"],
        )

    def test_parallel_completion_normalizes_legacy_duplicate_cards(self) -> None:
        cards, duplicate_count = deduplicate_cards([
            {"note_id": "n1", "title": "标题", "card_rank": 3},
            {"note_id": "n1", "card_cover_url": "https://img.example/n1.webp", "card_rank": 7},
            {"note_id": "n2", "title": "第二篇", "card_rank": 9},
        ])

        self.assertEqual(duplicate_count, 1)
        self.assertEqual([card["note_id"] for card in cards], ["n1", "n2"])
        self.assertEqual(cards[0]["card_rank"], 3)
        self.assertEqual(cards[0]["card_cover_url"], "https://img.example/n1.webp")

    def test_security_timeout_stops_new_access_and_preserves_checkpoint(self) -> None:
        class FakeLocator:
            def inner_text(self, **_kwargs):
                return "请完成安全验证后继续"

        class FakePage:
            def locator(self, _selector):
                return FakeLocator()

            def close(self):
                return None

        class FakeContext:
            def new_page(self):
                return FakePage()

        class FakePlaywright:
            def __enter__(self):
                return object()

            def __exit__(self, *_args):
                return False

        calls = []
        context = FakeContext()

        class FakeUpstream:
            @staticmethod
            def connect_browser(_playwright, _relay_port):
                return object()

            @staticmethod
            def get_or_create_context(_browser):
                return context

            @staticmethod
            def scrape_note(_page, card, **_kwargs):
                calls.append(card["note_id"])
                raise RuntimeError("请完成安全验证后继续")

        playwright_package = types.ModuleType("playwright")
        playwright_sync = types.ModuleType("playwright.sync_api")
        playwright_sync.sync_playwright = FakePlaywright

        with tempfile.TemporaryDirectory() as temporary:
            output = Path(temporary)
            cards = [
                {
                    "note_id": "n1",
                    "search_result_url": "https://example.test/search/n1?xsec_token=token",
                    "note_url": "https://example.test/explore/n1",
                },
                {"note_id": "n2", "note_url": "https://example.test/explore/n2"},
            ]
            (output / "xiaohongshu_cards_latest.json").write_text(
                json.dumps(cards, ensure_ascii=False),
                encoding="utf-8",
            )
            monotonic_values = itertools.chain(
                (0.0, 10.0, 20.0, 30.0),
                itertools.count(40.0, 10.0),
            )
            original_loader = body_completion.load_upstream
            original_monotonic = body_completion.time.monotonic
            original_sleep = body_completion.time.sleep
            try:
                body_completion.load_upstream = lambda _path: FakeUpstream()
                body_completion.time.monotonic = lambda: next(monotonic_values)
                body_completion.time.sleep = lambda _seconds: None
                with mock.patch.dict(
                    sys.modules,
                    {"playwright": playwright_package, "playwright.sync_api": playwright_sync},
                ):
                    summary = body_completion.complete_bodies(
                        output,
                        relay_port=18792,
                        workers=2,
                        attempts=3,
                        security_verification_timeout_seconds=5,
                    )
            finally:
                body_completion.load_upstream = original_loader
                body_completion.time.monotonic = original_monotonic
                body_completion.time.sleep = original_sleep

            self.assertEqual(summary["stopReason"], "security_verification_timeout")
            self.assertTrue(summary["newAccessStopped"])
            self.assertFalse(summary["transitionedToAnalysis"])
            self.assertEqual(summary["collectionStatus"], "partial")
            self.assertTrue(summary["workersExited"])
            self.assertEqual(summary["securityVerification"]["status"], "timed_out")
            self.assertEqual(summary["securityVerification"]["recoveryAction"], "manual_verification_then_resume")
            self.assertLessEqual(len(calls), 2)
            self.assertEqual(
                json.loads((output / "xiaohongshu_notes_latest.json").read_text(encoding="utf-8")),
                [],
            )

    def test_rate_limit_stops_immediately_and_preserves_partial_checkpoint(self) -> None:
        class FakeLocator:
            def inner_text(self, **_kwargs):
                return "安全限制 访问频繁，请稍后再试 300013"

        class FakePage:
            def locator(self, _selector):
                return FakeLocator()

            def close(self):
                return None

        class FakeContext:
            def new_page(self):
                return FakePage()

        class FakePlaywright:
            def __enter__(self):
                return object()

            def __exit__(self, *_args):
                return False

        calls = []
        context = FakeContext()

        class FakeUpstream:
            @staticmethod
            def connect_browser(_playwright, _relay_port):
                return object()

            @staticmethod
            def get_or_create_context(_browser):
                return context

            @staticmethod
            def scrape_note(_page, card, **_kwargs):
                calls.append(card["note_id"])
                return types.SimpleNamespace(
                    note_id=card["note_id"],
                    note_url=card["note_url"],
                    body="访问频繁，请稍后再试",
                    access_status="detail_rate_limited",
                )

        playwright_package = types.ModuleType("playwright")
        playwright_sync = types.ModuleType("playwright.sync_api")
        playwright_sync.sync_playwright = FakePlaywright

        with tempfile.TemporaryDirectory() as temporary:
            output = Path(temporary)
            cards = [
                {"note_id": "n1", "note_url": "https://example.test/explore/n1"},
                {"note_id": "n2", "note_url": "https://example.test/explore/n2"},
            ]
            (output / "xiaohongshu_cards_latest.json").write_text(json.dumps(cards), encoding="utf-8")
            original_loader = body_completion.load_upstream
            try:
                body_completion.load_upstream = lambda _path: FakeUpstream()
                with mock.patch.dict(
                    sys.modules,
                    {"playwright": playwright_package, "playwright.sync_api": playwright_sync},
                ):
                    started_at = time.perf_counter()
                    summary = body_completion.complete_bodies(
                        output,
                        relay_port=18792,
                        workers=1,
                        attempts=3,
                        rate_limit_auto_recovery=False,
                    )
                    elapsed_seconds = time.perf_counter() - started_at
            finally:
                body_completion.load_upstream = original_loader

            self.assertEqual(calls, ["n1"])
            self.assertEqual(summary["stopReason"], "rate_limited")
            self.assertFalse(summary["transitionedToAnalysis"])
            self.assertEqual(summary["collectionStatus"], "partial")
            self.assertTrue(summary["workersExited"])
            self.assertTrue(summary["newAccessStopped"])
            self.assertEqual(summary["rateLimit"]["status"], "stopped")
            self.assertEqual(summary["securityVerification"]["status"], "not_detected")
            self.assertLess(elapsed_seconds, 5.0)

    def test_default_candidate_profile_is_project_relative(self) -> None:
        self.assertEqual(DEFAULT_CANDIDATE_PROFILE, PROJECT_ROOT / "profiles/candidate_profile.json")
        self.assertEqual(resolve_project_path("profiles/candidate_profile.json"), DEFAULT_CANDIDATE_PROFILE.resolve())

    def test_stage_lines_follow_job_manager_contract(self) -> None:
        stream = io.StringIO()
        with redirect_stdout(stream):
            for index, label in enumerate(("coverage", "time", "memory", "info", "capabilities", "outreach", "review", "quality"), start=1):
                emit_stage(index, label)
        lines = stream.getvalue().splitlines()
        self.assertEqual(len(lines), 8)
        self.assertEqual(lines[0], "AGENT_STAGE 1/8 coverage completed")
        self.assertEqual(lines[-1], "AGENT_STAGE 8/8 quality completed")

    def test_workflow_summary_exposes_frontend_metrics(self) -> None:
        cards = [{"note_id": "n1", "title": "AI one"}, {"note_id": "n2", "title": "AI two"}]
        notes = [
            {
                "note_id": "n1",
                "title": "AI one",
                "body": "要求熟练使用Codex，邮箱 jobs@example.com",
                "publish_time": "昨天 20:05",
                "scraped_at": "2026-07-28T09:31:42",
            },
            {
                "note_id": "n2",
                "title": "AI two",
                "body": "需要AI项目经验，可以私信。",
                "publish_time": "2小时前",
                "scraped_at": "2026-07-28T09:31:42",
            },
        ]
        result = ApplicationIntelligencePipeline(PROFILE, now=COLLECTED).run(cards, notes)
        summary = build_workflow_summary(result.payload)
        self.assertEqual(summary["cardsDiscovered"], 2)
        self.assertEqual(summary["notesCollected"], 2)
        self.assertEqual(summary["bodiesCaptured"], 2)
        self.assertEqual(summary["bodyCoveragePercent"], 100.0)
        self.assertEqual(summary["timeNormalizedCount"], 2)
        self.assertEqual(summary["exactTimeCount"], 1)
        self.assertEqual(summary["estimatedTimeCount"], 1)
        self.assertGreaterEqual(summary["contactsFound"], 1)
        self.assertEqual(summary["greetingsGenerated"], 2)
        self.assertEqual(summary["emailsGenerated"], 2)
        self.assertEqual(summary["jobCardsGenerated"], 2)
        self.assertEqual(summary["applicationCopyGenerated"], 2)
        self.assertEqual(summary["generationCoveragePercent"], 100.0)
        self.assertEqual(summary["applicationInfo"], 2)
        self.assertEqual(summary["draftsGenerated"], 2)
        self.assertEqual(len(summary["agentStages"]), 8)

    def test_workflow_summary_does_not_misclassify_generic_partial_collection(self) -> None:
        cards = [{"note_id": "n1", "title": "AI one"}]
        result = ApplicationIntelligencePipeline(PROFILE, now=COLLECTED).run(cards, [])
        summary = build_workflow_summary(
            result.payload,
            {
                "collectionStatus": "partial",
                "missingBodies": 1,
                "transitionedToAnalysis": False,
                "stopReason": "scrape_runner_failed",
            },
        )

        self.assertEqual(summary["status"], "completed_partial")
        self.assertEqual(summary["analysisMode"], "partial_collection")
        self.assertEqual(summary["collectionStopReason"], "scrape_runner_failed")
        self.assertEqual(summary["securityVerification"], {})

    def test_job_ai_targets_parse_ready_bodies_without_losing_pending_resume_targets(self) -> None:
        cards = [
            {"note_id": "ready", "title": "AI 产品实习"},
            {"note_id": "pending", "title": "数据产品实习"},
        ]
        notes = [{
            "note_id": "ready",
            "title": "AI 产品实习",
            "body": "负责 AI 产品需求分析，要求熟悉 SQL。",
            "publish_time": "刚刚",
            "scraped_at": "2026-07-28T09:31:42",
        }]
        result = ApplicationIntelligencePipeline(PROFILE, now=COLLECTED).run(cards, notes)

        targets, reason, pending_count, fully_deferred = partition_job_ai_targets(
            result.payload,
            None,
            {"missingBodies": 1, "stopReason": "rate_limited"},
        )

        self.assertEqual(targets, {"ready"})
        self.assertEqual(reason, "rate_limited")
        self.assertEqual(pending_count, 1)
        self.assertFalse(fully_deferred)
        self.assertNotEqual(result.payload.get("codex_runtime", {}).get("status"), "deferred_missing_bodies")
        self.assertEqual(result.payload["source_coverage"]["readyCount"], 1)
        summary = build_workflow_summary(result.payload, {"missingBodies": 1, "stopReason": "rate_limited"})
        self.assertTrue(summary["qualityPending"])
        self.assertEqual(summary["sourceCoverage"]["pendingCount"], 1)
        self.assertEqual(summary["issues"][0]["code"], "SOURCE_BODY_COMPLETION_PENDING")
        self.assertEqual(summary["agentStages"][5]["status"], "partial")
        self.assertEqual(summary["agentStages"][6]["status"], "pending")
        self.assertEqual(summary["agentStages"][-1]["status"], "pending")

    def test_job_ai_targets_accept_url_only_records_by_canonical_note_id(self) -> None:
        payload = {
            "records": [{
                "note_url": "https://example.test/discovery/item/url-only?token=signed",
                "body": "Complete role description",
            }],
            "quality_gate": {"discovered_count": 1, "body_count": 1},
        }

        targets, reason, pending_count, fully_deferred = partition_job_ai_targets(
            payload,
            {"url-only"},
            {"missingBodies": 0},
        )

        self.assertEqual(targets, {"url-only"})
        self.assertEqual(reason, "")
        self.assertEqual(pending_count, 0)
        self.assertFalse(fully_deferred)

    def test_workflow_summary_marks_only_explicit_security_timeout(self) -> None:
        cards = [{"note_id": "n1", "title": "AI one"}]
        result = ApplicationIntelligencePipeline(PROFILE, now=COLLECTED).run(cards, [])
        summary = build_workflow_summary(
            result.payload,
            {
                "collectionStatus": "partial",
                "missingBodies": 1,
                "transitionedToAnalysis": False,
                "stopReason": "security_verification_timeout",
                "securityVerification": {"status": "timed_out", "timeoutSeconds": 90},
            },
        )

        self.assertEqual(summary["analysisMode"], "security_timeout_partial")
        self.assertEqual(summary["securityVerification"]["status"], "timed_out")

    def test_search_security_exit_materializes_partial_results_for_resume(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            output = Path(temporary) / "run" / "artifacts"
            output.mkdir(parents=True)
            profile = Path(temporary) / "profile.json"
            profile.write_text(json.dumps(PROFILE, ensure_ascii=False), encoding="utf-8")
            (output / "xiaohongshu_cards_latest.json").write_text(
                json.dumps(
                    [{"note_id": "n1", "note_url": "https://example.test/n1", "title": "AI 岗位"}],
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )
            (output / "security-restriction.json").write_text(
                json.dumps(
                    {
                        "status": "timed_out",
                        "phase": "search_discovery",
                        "timeoutSeconds": 600,
                        "recoveryAction": "manual_verification_then_resume",
                    }
                ),
                encoding="utf-8",
            )

            with mock.patch.object(workflow, "resolve_upstream_runner", return_value=Path("runner.py")):
                with mock.patch.object(
                    workflow.subprocess,
                    "run",
                    return_value=subprocess.CompletedProcess(["runner"], 3),
                ):
                    return_code = workflow.main(
                        [
                            "--output-dir",
                            str(output),
                            "--candidate-profile",
                            str(profile),
                        ]
                    )

            summary = json.loads((output / "workflow-summary.json").read_text(encoding="utf-8"))
            self.assertEqual(return_code, 3)
            self.assertEqual(summary["analysisMode"], "security_timeout_partial")
            self.assertEqual(summary["collectionStopReason"], "security_verification_timeout")
            self.assertEqual(summary["securityVerification"]["phase"], "search_discovery")
            self.assertEqual(summary["jobCardsGenerated"], 0)
            self.assertEqual(summary["applicationCopyGenerated"], 0)
            self.assertEqual(summary["sourceCoverage"]["targetCount"], 1)
            self.assertEqual(summary["sourceCoverage"]["readyCount"], 0)
            self.assertEqual(summary["sourceCoverage"]["pendingCount"], 1)

    def test_project_manifest_hashes_every_artifact_except_itself(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            output = Path(temporary)
            (output / "nested").mkdir()
            (output / "one.json").write_text('{"ok":true}', encoding="utf-8")
            (output / "nested" / "two.csv").write_text("a,b\n1,2\n", encoding="utf-8")
            (output / "workflow.stderr.log").write_text("", encoding="utf-8")
            (output / "artifact-manifest.json").write_text("stale", encoding="utf-8")
            summary = {
                "checks": {"coverage": True, "quality": True},
                "issues": [],
                "notesCollected": 2,
                "bodiesCaptured": 2,
            }
            manifest_path = write_project_manifest(output, summary)
            manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
            self.assertEqual(manifest["runner"], "xiaohongshu-project-workflow")
            self.assertEqual(manifest["status"], "succeeded")
            self.assertEqual({item["path"] for item in manifest["artifacts"]}, {"one.json", "nested/two.csv"})
            for item in manifest["artifacts"]:
                artifact = output / Path(item["path"])
                self.assertEqual(item["bytes"], artifact.stat().st_size)
                self.assertEqual(item["sha256"], hashlib.sha256(artifact.read_bytes()).hexdigest())


if __name__ == "__main__":
    unittest.main()
