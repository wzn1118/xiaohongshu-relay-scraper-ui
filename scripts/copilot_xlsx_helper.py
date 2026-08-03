"""Narrow openpyxl bridge used by the Data Copilot artifact service."""

from __future__ import annotations

import json
import os
import sys
from pathlib import Path


def load_openpyxl():
    try:
        import openpyxl  # type: ignore
    except ImportError:
        print(json.dumps({"available": False}), flush=True)
        raise SystemExit(3)
    return openpyxl


def safe_cell(value):
    if value is None:
        return ""
    if isinstance(value, (dict, list)):
        text = json.dumps(value, ensure_ascii=False, separators=(",", ":"))
    else:
        text = str(value)
    if text.lstrip("\t\r ").startswith(("=", "+", "-", "@")):
        return "'" + text
    return text


def create_workbook(openpyxl, target: Path) -> None:
    payload = json.load(sys.stdin)
    columns = payload.get("columns") or []
    rows = payload.get("rows") or []
    if not isinstance(columns, list) or not isinstance(rows, list):
        raise ValueError("columns and rows must be arrays")
    workbook = openpyxl.Workbook(write_only=True)
    worksheet = workbook.create_sheet(title=str(payload.get("sheetName") or "Data")[:31])
    default = workbook["Sheet"] if "Sheet" in workbook.sheetnames else None
    if default is not None:
        workbook.remove(default)
    worksheet.append([safe_cell(value) for value in columns])
    for row in rows:
        if not isinstance(row, list):
            raise ValueError("each row must be an array")
        worksheet.append([safe_cell(value) for value in row])
    target.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(target)


def inspect_workbook(openpyxl, source: Path, max_rows: int) -> dict:
    workbook = openpyxl.load_workbook(source, read_only=True, data_only=True)
    try:
        sheets = []
        for worksheet in workbook.worksheets:
            preview = []
            for index, row in enumerate(worksheet.iter_rows(values_only=True)):
                if index >= max_rows:
                    break
                preview.append([safe_cell(value) for value in row[:256]])
            sheets.append(
                {
                    "name": worksheet.title,
                    "maxRow": int(worksheet.max_row or 0),
                    "maxColumn": int(worksheet.max_column or 0),
                    "preview": preview,
                    "truncated": int(worksheet.max_row or 0) > max_rows
                    or int(worksheet.max_column or 0) > 256,
                }
            )
        return {
            "kind": "workbook",
            "sheetNames": workbook.sheetnames,
            "sheets": sheets,
            "truncated": any(sheet["truncated"] for sheet in sheets),
        }
    finally:
        workbook.close()


def main() -> int:
    if len(sys.argv) < 2:
        return 2
    command = sys.argv[1]
    openpyxl = load_openpyxl()
    if command == "probe":
        print(json.dumps({"available": True, "version": openpyxl.__version__}), flush=True)
        return 0
    if command == "create" and len(sys.argv) == 3:
        target = Path(sys.argv[2]).resolve()
        create_workbook(openpyxl, target)
        print(json.dumps({"created": True, "size": os.path.getsize(target)}), flush=True)
        return 0
    if command == "inspect" and len(sys.argv) == 4:
        source = Path(sys.argv[2]).resolve()
        max_rows = max(1, min(10_000, int(sys.argv[3])))
        print(json.dumps(inspect_workbook(openpyxl, source, max_rows), ensure_ascii=False), flush=True)
        return 0
    return 2


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as error:  # Error text is bounded and handled by Node.
        print(f"{type(error).__name__}: {error}", file=sys.stderr, flush=True)
        raise SystemExit(1)
