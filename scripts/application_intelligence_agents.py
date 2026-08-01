from __future__ import annotations

import csv
import json
import re
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta, timezone
from pathlib import Path
from typing import Any, Iterable

from codex_runtime_outreach import PROMPT_VERSION


SHANGHAI = timezone(timedelta(hours=8), name="Asia/Shanghai")
TIME_FIELDS = ("publish_time", "card_publish_time", "card_text_segments", "source_card_text", "card_author")
TEXT_FIELDS = ("body", "title", "source_card_text", "card_text_segments")
ILLEGAL_SHEET_CHARACTERS = re.compile(r"[\x00-\x08\x0B-\x0C\x0E-\x1F]")


def _text(value: Any) -> str:
    return value.strip() if isinstance(value, str) else ""


def _media_values(value: Any) -> list[str]:
    if isinstance(value, list):
        candidates = value
    elif isinstance(value, str):
        candidates = re.split(r"\s*\|\s*", value)
    else:
        candidates = []
    return [str(item).strip() for item in candidates if str(item).strip()]


def _is_content_image_url(value: str) -> bool:
    lowered = value.lower()
    return bool(re.match(r"^https?://", value)) and not any(
        marker in lowered for marker in ("sns-avatar", "/avatar/", "avatar_")
    )


def build_media(note: dict[str, Any]) -> dict[str, Any]:
    sources = [
        ("detail", _media_values(note.get("detail_image_urls")), _media_values(note.get("detail_image_alts"))),
        ("card", _media_values(note.get("card_image_urls")), []),
        ("cover", _media_values(note.get("card_cover_url")), [_text(note.get("card_cover_alt"))]),
    ]
    images: list[dict[str, str]] = []
    seen: set[str] = set()
    alt_texts: list[str] = []
    for source, urls, alts in sources:
        for index, url in enumerate(urls):
            if url in seen or not _is_content_image_url(url):
                continue
            seen.add(url)
            alt = alts[index] if index < len(alts) else ""
            images.append({"url": url, "alt": alt, "source": source})
            if alt and alt not in alt_texts:
                alt_texts.append(alt)
    if images and alt_texts:
        analysis = {
            "status": "alt_text_available",
            "summary": " | ".join(alt_texts[:6]),
            "job_signals": [],
            "source": "image_alt_text",
        }
    elif images:
        analysis = {
            "status": "pending_ai",
            "summary": "",
            "job_signals": [],
            "source": "image_urls",
        }
    else:
        analysis = {
            "status": "no_images",
            "summary": "",
            "job_signals": [],
            "source": "none",
        }
    return {
        "cover_url": _text(note.get("card_cover_url")) or (images[0]["url"] if images else ""),
        "images": images,
        "analysis": analysis,
    }


def _sheet_safe(value: Any) -> Any:
    if isinstance(value, str):
        value = ILLEGAL_SHEET_CHARACTERS.sub("", value)
        if value.startswith(("=", "+", "-", "@")):
            return "'" + value
    return value


def _aware_datetime(value: Any, fallback: datetime) -> tuple[datetime, str]:
    raw = _text(value)
    if not raw:
        return fallback, "pipeline_fallback"
    try:
        parsed = datetime.fromisoformat(raw.replace("Z", "+00:00"))
        if parsed.tzinfo is None:
            parsed = parsed.replace(tzinfo=SHANGHAI)
        return parsed.astimezone(SHANGHAI), "scraped_at"
    except ValueError:
        return fallback, "pipeline_fallback_invalid_scraped_at"


def _iso_minute(value: datetime) -> str:
    return value.replace(second=0, microsecond=0).isoformat(timespec="minutes")


def _time_source(note: dict[str, Any]) -> tuple[str, str]:
    marker = re.compile(
        r"(?:今天|昨天|前天|\d+\s*(?:天|小时|分钟)前|"
        r"\d{4}\s*[-/.年]\s*\d{1,2}\s*[-/.月]\s*\d{1,2}|"
        r"\d{1,2}\s*[-/.月]\s*\d{1,2})"
    )
    for field in TIME_FIELDS:
        value = _text(note.get(field))
        if not value:
            continue
        if field in {"publish_time", "card_publish_time"} or marker.search(value):
            match = marker.search(value)
            if field not in {"publish_time", "card_publish_time"} and match:
                value = value[match.start() :]
            return value, field
    return "", ""


def normalize_publish_time(raw: str, collected_at: datetime, source_field: str = "publish_time") -> dict[str, Any]:
    value = re.sub(r"\s+", " ", raw.strip())
    value = re.sub(r"\s+(?:北京|上海|广东|浙江|江苏|四川|湖北|湖南|海外)$", "", value)
    result: dict[str, Any] = {
        "raw": raw,
        "value": "",
        "timezone": "Asia/Shanghai",
        "precision": "unknown",
        "method": "unparsed",
        "is_estimated": False,
        "source_field": source_field,
        "collected_at": collected_at.isoformat(timespec="seconds"),
    }
    if not value:
        result["method"] = "missing"
        return result

    relative_named = re.search(r"(?P<label>今天|昨天|前天)(?:\s+(?P<hour>\d{1,2}):(?P<minute>\d{2}))?", value)
    if relative_named:
        days = {"今天": 0, "昨天": 1, "前天": 2}[relative_named.group("label")]
        target_date = collected_at.date() - timedelta(days=days)
        if relative_named.group("hour") is not None:
            target = datetime.combine(
                target_date,
                time(int(relative_named.group("hour")), int(relative_named.group("minute"))),
                tzinfo=SHANGHAI,
            )
            result.update(
                value=_iso_minute(target),
                precision="minute",
                method="relative_named_with_clock",
                is_estimated=False,
            )
        else:
            result.update(
                value=target_date.isoformat(),
                precision="day",
                method="relative_named_date_only",
                is_estimated=True,
            )
        return result

    relative = re.search(r"(?P<count>\d+)\s*(?P<unit>天|小时|分钟)前", value)
    if relative:
        count = int(relative.group("count"))
        unit = relative.group("unit")
        if unit == "天":
            target_date = collected_at.date() - timedelta(days=count)
            result.update(
                value=target_date.isoformat(),
                precision="day",
                method="relative_days_date_only",
                is_estimated=True,
            )
            return result
        delta = {
            "小时": timedelta(hours=count),
            "分钟": timedelta(minutes=count),
        }[unit]
        target = collected_at - delta
        result.update(
            value=_iso_minute(target),
            precision={"小时": "relative_hour", "分钟": "relative_minute"}[unit],
            method=f"relative_{unit}",
            is_estimated=True,
        )
        return result

    absolute = re.search(
        r"(?:(?P<year>\d{4})\s*[-/.年]\s*)?"
        r"(?P<month>\d{1,2})\s*[-/.月]\s*(?P<day>\d{1,2})(?:日)?"
        r"(?:\s+(?P<hour>\d{1,2}):(?P<minute>\d{2}))?",
        value,
    )
    if absolute:
        year = int(absolute.group("year") or collected_at.year)
        try:
            target_date = date(year, int(absolute.group("month")), int(absolute.group("day")))
        except ValueError:
            result["method"] = "invalid_absolute_date"
            return result
        if absolute.group("year") is None and target_date > collected_at.date():
            target_date = date(year - 1, target_date.month, target_date.day)
        if absolute.group("hour") is None:
            result.update(
                value=target_date.isoformat(),
                precision="day",
                method="absolute_date",
                is_estimated=False,
            )
            return result
        target = datetime.combine(
            target_date,
            time(int(absolute.group("hour")), int(absolute.group("minute"))),
            tzinfo=SHANGHAI,
        )
        result.update(
            value=_iso_minute(target),
            precision="minute",
            method="absolute_datetime",
            is_estimated=False,
        )
        return result

    return result


def _evidence_excerpt(text: str, start: int, end: int, radius: int = 45) -> str:
    return text[max(0, start - radius) : min(len(text), end + radius)].strip()


def _provenance_item(kind: str, value: str, field: str, text: str, start: int, end: int) -> dict[str, Any]:
    return {
        "type": kind,
        "value": value,
        "source_field": field,
        "evidence": _evidence_excerpt(text, start, end),
        "offset_start": start,
        "offset_end": end,
    }


class ApplicationInfoAgent:
    EMAIL_RE = re.compile(r"(?<![\w.+-])[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}(?![\w.-])", re.I)
    WECHAT_RE = re.compile(r"(?:微信|vx|v信|wechat)\s*(?:号|ID)?\s*[:：]?\s*([A-Za-z][A-Za-z0-9_-]{5,19})", re.I)
    PHONE_RE = re.compile(r"(?:电话|手机|联系)\s*[:：]?\s*(1[3-9]\d{9})(?!\d)")
    URL_RE = re.compile(r"https?://[^\s|，。；;]+", re.I)
    REQUIREMENT_CUES = (
        "任职要求",
        "岗位要求",
        "要求",
        "适合",
        "希望",
        "需要",
        "优先",
        "熟练",
        "擅长",
        "英语",
        "每周",
        "到岗",
        "实习",
        "base",
        "经验",
        "专业",
        "学历",
        "能力",
    )
    RESPONSIBILITY_CUES = (
        "岗位职责",
        "工作职责",
        "职位描述",
        "岗位描述",
        "职责描述",
        "工作内容",
        "职责",
        "负责",
        "协助",
        "参与",
        "支持",
        "跟进",
        "维护",
        "运营",
        "策划",
        "分析",
        "撰写",
        "对接",
    )
    REQUIREMENT_WEIGHTS = {
        "任职要求": 4,
        "岗位要求": 4,
        "要求": 2,
        "需要": 2,
        "优先": 2,
        "熟练": 2,
        "擅长": 2,
        "每周": 2,
        "到岗": 2,
        "经验": 2,
        "专业": 2,
        "学历": 2,
        "能力": 2,
    }
    RESPONSIBILITY_WEIGHTS = {
        "岗位职责": 4,
        "工作职责": 4,
        "职位描述": 4,
        "岗位描述": 4,
        "职责描述": 4,
        "工作内容": 4,
        "职责": 2,
        "负责": 2,
        "协助": 2,
        "参与": 2,
        "支持": 2,
        "跟进": 2,
        "维护": 2,
        "策划": 2,
        "分析": 2,
        "撰写": 2,
        "对接": 2,
    }
    SECTION_CATEGORIES = {
        "岗位职责": "responsibility",
        "工作职责": "responsibility",
        "职位描述": "responsibility",
        "岗位描述": "responsibility",
        "职责描述": "responsibility",
        "工作内容": "responsibility",
        "任职要求": "requirement",
        "岗位要求": "requirement",
        "职位要求": "requirement",
        "任职资格": "requirement",
        "任职条件": "requirement",
        "职位条件": "requirement",
    }
    SECTION_HEADING_RE = re.compile(
        "|".join(re.escape(value) for value in sorted(SECTION_CATEGORIES, key=len, reverse=True)),
        re.I,
    )
    LIST_MARKER_RE = re.compile(
        r"(?:^|(?<=[\s:：。；;]))"
        r"(?:\d{1,2}\s*[、.．)）]|[一二三四五六七八九十]{1,3}\s*[、.．)）]|"
        r"[（(]\s*(?:\d{1,2}|[一二三四五六七八九十]{1,3})\s*[）)]|[•●▪◦·])\s*"
    )

    @classmethod
    def _section_spans(cls, body: str) -> list[tuple[int, int, str | None, str | None]]:
        headings = list(cls.SECTION_HEADING_RE.finditer(body))
        if not headings:
            return [(0, len(body), None, None)]

        spans: list[tuple[int, int, str | None, str | None]] = []
        if headings[0].start() > 0:
            spans.append((0, headings[0].start(), None, None))
        for index, heading in enumerate(headings):
            end = headings[index + 1].start() if index + 1 < len(headings) else len(body)
            label = heading.group(0)
            spans.append((heading.end(), end, cls.SECTION_CATEGORIES[label], label))
        return spans

    @classmethod
    def _list_item_spans(cls, body: str, start: int, end: int) -> list[tuple[int, int]]:
        section = body[start:end]
        markers = list(cls.LIST_MARKER_RE.finditer(section))
        if not markers:
            return [(start, end)]

        spans: list[tuple[int, int]] = []
        if section[: markers[0].start()].strip(" \t\r\n:：-—"):
            spans.append((start, start + markers[0].start()))
        for index, marker in enumerate(markers):
            item_start = start + marker.end()
            item_end = start + (markers[index + 1].start() if index + 1 < len(markers) else len(section))
            spans.append((item_start, item_end))
        return spans

    @classmethod
    def _classification_clauses(cls, body: str) -> Iterable[tuple[int, str, str | None, str | None]]:
        boundary_cues = tuple(dict.fromkeys(cls.REQUIREMENT_CUES + cls.RESPONSIBILITY_CUES))
        boundary_pattern = re.compile(rf"[，,]\s*(?=(?:{'|'.join(map(re.escape, boundary_cues))}))", re.I)
        for section_start, section_end, section_category, section_label in cls._section_spans(body):
            for item_start, item_end in cls._list_item_spans(body, section_start, section_end):
                item_text = body[item_start:item_end]
                for match in re.finditer(r"[^。！？；;\n]+[。！？；;]?", item_text):
                    raw_sentence = match.group(0)
                    starts = [0, *(boundary.end() for boundary in boundary_pattern.finditer(raw_sentence))]
                    ends = [*starts[1:], len(raw_sentence)]
                    for local_start, local_end in zip(starts, ends):
                        raw_clause = raw_sentence[local_start:local_end]
                        leading = len(raw_clause) - len(raw_clause.lstrip(" \t\r\n:：-—"))
                        sentence = raw_clause.strip(" \t\r\n:：-—")
                        if sentence:
                            yield (
                                item_start + match.start() + local_start + leading,
                                sentence,
                                section_category,
                                section_label,
                            )

    def run(self, note: dict[str, Any]) -> dict[str, Any]:
        contacts: list[dict[str, Any]] = []
        seen: set[tuple[str, str]] = set()
        routes: list[dict[str, Any]] = []

        for field in TEXT_FIELDS:
            text = _text(note.get(field))
            if not text:
                continue
            patterns: tuple[tuple[str, re.Pattern[str], int], ...] = (
                ("email", self.EMAIL_RE, 0),
                ("wechat", self.WECHAT_RE, 1),
                ("phone", self.PHONE_RE, 1),
                ("url", self.URL_RE, 0),
            )
            for kind, pattern, group in patterns:
                for match in pattern.finditer(text):
                    found = match.group(group)
                    key = (kind, found.lower())
                    if key in seen:
                        continue
                    seen.add(key)
                    contacts.append(_provenance_item(kind, found, field, text, match.start(group), match.end(group)))
            for label, route_pattern in (
                ("xiaohongshu_dm", r"(?:私信|私我|评论区|评论留言)"),
                ("email", r"(?:邮箱|邮件|简历投递)"),
                ("official_site", r"(?:官网|招聘网站|招聘系统)"),
            ):
                match = re.search(route_pattern, text, re.I)
                if match and not any(item["type"] == label for item in routes):
                    routes.append(_provenance_item(label, label, field, text, match.start(), match.end()))

        requirements: list[dict[str, Any]] = []
        responsibilities: list[dict[str, Any]] = []
        source_field = "body"
        body = _text(note.get(source_field))
        if not body:
            for fallback_field in ("source_card_text", "card_text_segments", "title", "card_title"):
                body = _text(note.get(fallback_field))
                if body:
                    source_field = fallback_field
                    break
        classified: set[str] = set()
        for offset_start, sentence, section_category, section_label in self._classification_clauses(body):
            lowered = sentence.lower()
            if len(sentence) < 4:
                continue
            normalized = re.sub(r"[\W_]+", "", lowered)
            if not normalized or normalized in classified:
                continue
            requirement_hits = [cue for cue in self.REQUIREMENT_CUES if cue in lowered]
            responsibility_hits = [cue for cue in self.RESPONSIBILITY_CUES if cue in lowered]
            requirement_score = sum(self.REQUIREMENT_WEIGHTS.get(cue, 1) for cue in requirement_hits)
            responsibility_score = sum(self.RESPONSIBILITY_WEIGHTS.get(cue, 1) for cue in responsibility_hits)
            if not section_category and max(requirement_score, responsibility_score) < 2:
                continue
            if section_category:
                category = section_category
                if (
                    section_category == "responsibility"
                    and requirement_score >= 4
                    and requirement_score > responsibility_score
                ):
                    category = "requirement"
            elif requirement_score == responsibility_score:
                if any(cue in lowered for cue in ("任职要求", "岗位要求")):
                    category = "requirement"
                elif any(cue in lowered for cue in ("岗位职责", "工作职责", "工作内容")):
                    category = "responsibility"
                else:
                    continue
            else:
                category = "requirement" if requirement_score > responsibility_score else "responsibility"
            winner = max(requirement_score, responsibility_score)
            loser = min(requirement_score, responsibility_score)
            confidence = (
                0.96
                if section_category and winner >= 2
                else 0.9
                if section_category
                else round(min(0.98, 0.55 + winner * 0.07 - loser * 0.03), 2)
            )
            classification_basis = requirement_hits if category == "requirement" else responsibility_hits
            if section_label and category == section_category:
                classification_basis = [section_label, *classification_basis]
            item = {
                "text": sentence,
                "source_field": source_field,
                "evidence": sentence,
                "offset_start": offset_start,
                "offset_end": offset_start + len(sentence),
                "classification_confidence": confidence,
                "classification_basis": list(dict.fromkeys(classification_basis)),
            }
            classified.add(normalized)
            if category == "requirement":
                requirements.append(item)
            else:
                responsibilities.append(item)

        return {
            "contacts": contacts,
            "application_routes": routes,
            "responsibilities": responsibilities,
            "requirements": requirements,
        }


def _string_values(value: Any) -> Iterable[str]:
    if isinstance(value, str) and value.strip():
        yield value.strip()
    elif isinstance(value, list):
        for item in value:
            yield from _string_values(item)
    elif isinstance(value, dict):
        for item in value.values():
            yield from _string_values(item)


def _candidate_application_profile(profile: dict[str, Any]) -> dict[str, str]:
    for key in ("candidate_application", "candidateProfile"):
        value = profile.get(key)
        if isinstance(value, dict):
            return {
                field: _text(value.get(field))
                for field in (
                    "name",
                    "school",
                    "major",
                    "degreeYear",
                    "phoneWeChat",
                    "email",
                    "availabilityDays",
                    "internshipDuration",
                )
            }
    return {}


def _candidate_name(profile: dict[str, Any]) -> str:
    application_profile = _candidate_application_profile(profile)
    if application_profile.get("name"):
        return application_profile["name"]
    for key in ("name", "display_name", "chinese_name", "candidate_name"):
        value = _text(profile.get(key))
        if value:
            return value
    candidate = profile.get("candidate")
    if isinstance(candidate, dict):
        return _candidate_name(candidate)
    return "候选人"


def load_candidate_evidence(profile: dict[str, Any]) -> list[dict[str, Any]]:
    explicit = profile.get("evidence_items")
    if isinstance(explicit, list):
        result = []
        for index, item in enumerate(explicit, start=1):
            if not isinstance(item, dict):
                continue
            raw_confidence = item.get("confidence")
            try:
                confidence = int(float(str(raw_confidence))) if raw_confidence is not None else None
            except (TypeError, ValueError):
                confidence = 0
            if confidence is not None and confidence < 75:
                continue
            detail = _text(item.get("detail") or item.get("description") or item.get("text"))
            if not detail:
                continue
            result.append(
                {
                    "id": _text(item.get("id")) or f"evidence-{index}",
                    "category": _text(item.get("category")) or "evidence",
                    "label": _text(item.get("label") or item.get("title")) or detail[:36],
                    "detail": detail,
                    "first_person_claim": _text(item.get("first_person_claim")),
                    "organization": _text(item.get("organization")),
                    "period": _text(item.get("period")),
                    "skills": item.get("skills") if isinstance(item.get("skills"), list) else [],
                    "outcomes": item.get("outcomes") if isinstance(item.get("outcomes"), list) else [],
                    "source": _text(item.get("source") or item.get("url")),
                    "source_evidence": _text(item.get("evidence")),
                    "confidence": confidence if confidence is not None else 0,
                }
            )
        return result

    result: list[dict[str, Any]] = []
    sections = ("education", "experience", "experiences", "projects", "github_projects", "skills")
    for section in sections:
        value = profile.get(section)
        if value is None:
            continue
        items = value if isinstance(value, list) else [value]
        for item in items:
            if isinstance(item, dict):
                label = next((_text(item.get(key)) for key in ("title", "name", "role", "company", "project") if _text(item.get(key))), section)
                details = [text for text in _string_values(item) if text != label]
                detail = "；".join(dict.fromkeys(details))
                source = next((_text(item.get(key)) for key in ("source", "source_pdf", "url", "github_url") if _text(item.get(key))), "")
            else:
                label = section
                detail = _text(item)
                source = ""
            if not detail:
                continue
            result.append(
                {
                    "id": _text(item.get("id")) if isinstance(item, dict) and _text(item.get("id")) else f"{section}-{len(result) + 1}",
                    "category": section,
                    "label": label,
                    "detail": detail,
                    "source": source,
                }
            )
    return result


MATCH_TERMS = (
    "ai",
    "codex",
    "python",
    "sql",
    "excel",
    "数据分析",
    "用户运营",
    "内容运营",
    "产品运营",
    "品牌",
    "市场",
    "营销",
    "社群",
    "增长",
    "海外",
    "英语",
    "github",
    "项目",
    "调研",
    "kol",
    "小红书",
    "自动化",
    "agent",
    "社交媒体",
    "活动",
    "文案",
)
GENERIC_MATCH_TERMS = {"项目", "活动", "运营"}
TOKEN_STOPWORDS = {
    "and",
    "for",
    "from",
    "intern",
    "internship",
    "role",
    "work",
    "working",
}
ROLE_MATCH_EXPANSIONS = (
    (re.compile(r"(?:商业分析|业务分析|经营分析|咨询)", re.I), "数据分析 市场调研 用户需求 竞品分析 研究报告 指标"),
    (re.compile(r"(?:市场研究|市场分析)", re.I), "市场调研 用户需求 竞品分析 数据分析 研究报告"),
    (re.compile(r"(?:品牌|公关)", re.I), "品牌 市场 舆情监测 分析 报告"),
    (re.compile(r"(?:内容运营|社媒运营)", re.I), "内容运营 社群运营 用户反馈 数据分析"),
)
NON_NARRATIVE_EVIDENCE_CATEGORIES = {"skills", "education"}
OUTREACH_FORMAT_VERSION = "fixed-cn-application-v1"


def _expand_role_target(target: str) -> str:
    expansions = [terms for pattern, terms in ROLE_MATCH_EXPANSIONS if pattern.search(target)]
    return " ".join([target, *expansions])


def _is_writable_evidence(item: dict[str, Any]) -> bool:
    category = _text(item.get("category")).lower()
    detail = _text(item.get("detail"))
    first_person_claim = _text(item.get("first_person_claim"))
    verified_v2 = bool(first_person_claim.startswith("我") and _text(item.get("source_evidence")))
    minimum_length = 8 if verified_v2 else 20
    if category in NON_NARRATIVE_EVIDENCE_CATEGORIES or len(detail) < minimum_length:
        return False
    return bool(
        re.search(
            r"(?:负责|搭建|分析|整理|优化|推进|开展|撰写|输出|设计|协同|完成|支持|支撑|运营|监测|调研|抓取|策划|对接|访谈|梳理|促成)",
            f"{detail} {first_person_claim}",
        )
    )


def _match_metrics(target: str, evidence: dict[str, str]) -> dict[str, Any]:
    target_lower = target.lower()
    evidence_lower = f"{evidence['label']} {evidence['detail']}".lower()
    matched_terms = [term for term in MATCH_TERMS if term in target_lower and term in evidence_lower]
    specific_terms = [term for term in matched_terms if term not in GENERIC_MATCH_TERMS]
    generic_terms = [term for term in matched_terms if term in GENERIC_MATCH_TERMS]
    target_tokens = {
        token
        for token in re.findall(r"[a-z][a-z0-9+#.-]{1,}", target_lower, re.I)
        if token not in TOKEN_STOPWORDS
    }
    evidence_tokens = {
        token
        for token in re.findall(r"[a-z][a-z0-9+#.-]{1,}", evidence_lower, re.I)
        if token not in TOKEN_STOPWORDS
    }
    token_overlap = sorted(target_tokens & evidence_tokens)
    accepted = bool(specific_terms or len(token_overlap) >= 2)
    score = len(specific_terms) * 5 + min(len(generic_terms), 2) + min(len(token_overlap), 4) * 2
    confidence = 0.0 if not accepted else round(
        min(0.98, 0.58 + min(len(specific_terms), 3) * 0.1 + min(len(token_overlap), 3) * 0.05),
        2,
    )
    return {
        "accepted": accepted,
        "score": score,
        "confidence": confidence,
        "matched_terms": list(dict.fromkeys([*specific_terms, *generic_terms, *token_overlap])),
    }


def _match_score(target: str, evidence: dict[str, str]) -> int:
    return int(_match_metrics(target, evidence)["score"])


class FitEvidenceAgent:
    def __init__(self, profile: dict[str, Any]):
        self.profile = profile
        self.evidence = load_candidate_evidence(profile)

    def run(self, note: dict[str, Any], requirements: list[dict[str, Any]]) -> list[dict[str, Any]]:
        job_card = note.get("job_card") if isinstance(note.get("job_card"), dict) else {}
        target = _expand_role_target(" ".join(
            [_text(note.get("title")), _text(job_card.get("role_name")), _text(note.get("body"))]
            + [_text(item.get("text")) for item in requirements]
        ))
        ranked = sorted(
            ((_match_metrics(target, item), item) for item in self.evidence),
            key=lambda pair: (
                not _is_writable_evidence(pair[1]),
                -pair[0]["score"],
                -pair[0]["confidence"],
                pair[1]["id"],
            ),
        )
        accepted = [
            dict(
                item,
                match_score=metrics["score"],
                match_confidence=metrics["confidence"],
                matched_terms=metrics["matched_terms"],
                match_basis="validated_term_match",
            )
            for metrics, item in ranked
            if metrics["accepted"]
        ]
        narrative = [item for item in accepted if _is_writable_evidence(item)]
        return (narrative or accepted)[:3]


class OutreachWriterAgent:
    def __init__(self, profile: dict[str, Any]):
        self.profile = profile
        self.name = _candidate_name(profile)
        self.application_profile = _candidate_application_profile(profile)

    def _profile_sentences(self) -> list[str]:
        profile = self.application_profile
        sentences: list[str] = []
        education = "".join(
            part
            for part in (
                profile.get("school", ""),
                profile.get("major", ""),
                profile.get("degreeYear", ""),
            )
            if part
        )
        if education:
            sentences.append(f"我目前就读于{education}。")
        availability = profile.get("availabilityDays", "")
        duration = profile.get("internshipDuration", "")
        if availability and duration:
            sentences.append(f"我每周可实习{availability}天，预计可连续实习{duration}。")
        elif availability:
            sentences.append(f"我每周可实习{availability}天。")
        elif duration:
            sentences.append(f"我预计可连续实习{duration}。")
        return sentences

    @staticmethod
    def _evidence_sentence(item: dict[str, Any], target: str = "") -> str:
        detail = re.sub(
            r"^(?:多份|三份)?简历(?:共同)?(?:确认|显示|记载)的?[：:]?\s*",
            "",
            _text(item.get("detail")),
        ).strip()
        clauses = [clause.strip(" -，。；") for clause in re.split(r"[；\n]+", detail) if clause.strip(" -，。；")]
        clauses = [
            clause
            for clause in clauses
            if not re.fullmatch(r"(?:exp|proj)-[A-Za-z0-9-]+", clause, re.I)
            and not re.fullmatch(r"\d{4}年\d{1,2}月[–—-]\d{4}年\d{1,2}月", clause)
            and not re.fullmatch(r"https?://\S+", clause, re.I)
        ]
        concepts = (
            "数据", "分析", "监测", "抓取", "清洗", "看板", "指标", "用户", "市场", "竞品",
            "内容", "社群", "增长", "转化", "活动", "调研", "运营", "协作", "项目", "报告",
            "kol", "玩家", "反馈", "twitter", "工具",
        )
        target_lower = target.lower()

        def relevance(clause: str, context: str = target_lower) -> int:
            lowered = clause.lower()
            score = sum(6 for term in MATCH_TERMS if term in context and term in lowered)
            score += sum(2 for term in concepts if term in context and term in lowered)
            if "数据分析" in context:
                score += 12 if "数据" in lowered else 0
                score += 8 if "分析" in lowered else 0
                score += 4 if re.search(r"监测|抓取|清洗|看板|指标|转化", lowered) else 0
            context_tokens = set(re.findall(r"[a-z][a-z0-9+#.-]{1,}", context, re.I))
            clause_tokens = set(re.findall(r"[a-z][a-z0-9+#.-]{1,}", lowered, re.I))
            return score + min(len(context_tokens & clause_tokens), 5)

        action_candidates = [
            clause for clause in clauses
            if 8 <= len(clause) <= 110
            and re.search(r"(?:负责|搭建|分析|整理|优化|推进|开展|撰写|输出|设计|协同|完成|支持|支撑|运营|监测|调研|抓取)", clause)
        ]
        action = max(action_candidates, key=lambda clause: relevance(clause), default="")
        result_candidates = [
            clause for clause in clauses
            if clause != action and 6 <= len(clause) <= 90
            and re.search(r"\d|(?:提升|增长|缩短|降低|减少|达到|促成|沉淀|避免)", clause)
        ]
        result_context = f"{target_lower} {action.lower()}"
        result = max(result_candidates, key=lambda clause: relevance(clause, result_context), default="")
        common_length = 0
        for left, right in zip(action, result):
            if left != right:
                break
            common_length += 1
        if result and common_length >= 4:
            result = result[common_length:].lstrip("，、； ")
        summary = "；".join(filter(None, (action, result)))
        if not summary:
            claim = _text(item.get("first_person_claim")).strip("，。； ")
            if claim.startswith("我"):
                return re.sub(r"^我(?:曾经|曾)?", "", claim).strip()
        return summary.rstrip("，。；")

    @staticmethod
    def _evidence_lead(item: dict[str, Any]) -> str:
        label = _text(item.get("label")) or "相关实践"
        if label.lower() in {"skills", "education", "experience", "experiences", "projects", "evidence"}:
            label = "相关项目"
        if label.endswith("实习生"):
            label = label[:-1]
        if any(marker in label for marker in ("实习", "项目", "经历", "工作")):
            return f"在{label}期间，我"
        return f"在{label}相关实践中，我"

    @staticmethod
    def _work_method(focus: str) -> str:
        if any(term in focus for term in ("分析", "调研", "研究", "咨询", "结论输出")):
            return "我会先确认业务问题和判断口径，再整理关键事实、核验结论，并把结果转成可执行的交付物"
        if "内容运营" in focus or "文案" in focus:
            return "我会先明确目标受众和内容目标，再结合反馈与表现数据调整选题、表达和发布节奏"
        if "用户运营" in focus or "社群" in focus:
            return "我会先梳理用户分层和关键需求，再通过反馈与行为表现验证运营动作并持续复盘"
        if "增长" in focus:
            return "我会先拆解转化路径和关键指标，再用用户反馈与结果数据判断优先级并推进验证"
        return "我会先对齐目标和交付标准，再用事实验证判断，并将结论转化为可执行、可复盘的动作"

    @staticmethod
    def _role_focus(title: str, application_info: dict[str, Any], fit_evidence: list[dict[str, Any]]) -> str:
        role_source = " ".join([
            title,
            *[
                _text(item.get("text"))
                for field in ("responsibilities", "requirements")
                for item in application_info.get(field, [])
                if isinstance(item, dict)
            ],
        ]).lower()
        evidence_source = " ".join(
            f"{_text(item.get('label'))} {_text(item.get('detail'))}"
            for item in fit_evidence
            if isinstance(item, dict)
        ).lower()
        matched = [term for term in MATCH_TERMS if term in role_source and term in evidence_source]
        specific = [term for term in matched if term not in {"活动", "市场", "运营", "文案", "项目"}]
        if specific:
            return "、".join(dict.fromkeys(specific[:2]))
        if re.search(r"商业分析|业务分析|经营分析|咨询", role_source):
            return "信息分析与结论输出"
        if re.search(r"市场研究|市场分析|调研", role_source):
            return "市场研究与判断"
        generic = [term for term in matched if term not in {"项目", "活动"}]
        if generic:
            return "、".join(dict.fromkeys(generic[:2]))
        return "信息整理与任务交付"

    def run(
        self,
        note: dict[str, Any],
        application_info: dict[str, Any],
        fit_evidence: list[dict[str, Any]],
    ) -> dict[str, Any]:
        job_card = note.get("job_card") if isinstance(note.get("job_card"), dict) else {}
        title = _text(job_card.get("role_name")) or _text(note.get("title")) or "实习岗位"
        salutation = "您好"
        name = "" if self.name == "候选人" else self.name
        focus = self._role_focus(title, application_info, fit_evidence)
        availability = self.application_profile.get("availabilityDays", "")
        duration = self.application_profile.get("internshipDuration", "")
        availability_text = f"每周可实习{availability}天" if availability else ""
        writable_evidence = [item for item in fit_evidence if _is_writable_evidence(item)]
        scored_evidence = [
            item for item in writable_evidence
            if isinstance(item.get("match_score"), (int, float))
        ]
        if scored_evidence:
            top_score = max(int(item["match_score"]) for item in scored_evidence)
            score_floor = max(5, (top_score * 2 + 2) // 3)
            writable_evidence = [
                item for item in scored_evidence
                if int(item["match_score"]) >= score_floor
            ]
        writing_evidence = writable_evidence[:2]
        evidence_ids = [item["id"] for item in writing_evidence]
        identity = "".join(
            part for part in (
                self.application_profile.get("school", ""),
                self.application_profile.get("major", ""),
                self.application_profile.get("degreeYear", ""),
            ) if part
        )
        identity_text = f"，目前就读于{identity}" if identity else ""
        availability_parts = [
            availability_text,
            f"可连续实习{duration}" if duration else "",
        ]
        availability_sentence = "、".join(part for part in availability_parts if part)
        role_target = " ".join([
            title,
            *[
                _text(item.get("text"))
                for field in ("responsibilities", "requirements")
                for item in application_info.get(field, [])
                if isinstance(item, dict)
            ],
        ])
        subject = "｜".join(filter(None, (f"应聘{title}", name, f"每周可实习{availability}天" if availability else "")))
        if writing_evidence:
            evidence_preview = self._evidence_sentence(writing_evidence[0], role_target)
            if len(evidence_preview) > 58:
                evidence_preview = evidence_preview[:58].rstrip("，。；")
            greeting = (
                f"{salutation}，我是{name or '应聘者'}{identity_text}，想应聘「{title}」。"
                f"我曾{evidence_preview or '参与相关项目实践'}，这段经历与岗位所需的{focus}直接相关。"
                f"{availability_sentence + '。' if availability_sentence else ''}请问岗位目前是否仍在招聘？"
            )
        else:
            greeting = (
                f"{salutation}，我是{name or '应聘者'}{identity_text}，想应聘「{title}」。"
                f"{availability_sentence + '，到岗安排稳定。' if availability_sentence else ''}"
                "请问岗位目前是否仍在招聘？期待进一步沟通，谢谢。"
            )
        if not writing_evidence:
            email_body = (
                "尊敬的招聘负责人：\n"
                f"您好！我是{name or '应聘者'}{identity_text}，了解到贵公司的「{title}」岗位后，希望申请该职位。\n\n"
                "我希望进一步了解岗位负责的业务方向、核心任务与交付要求，并据此补充最相关的项目案例。\n\n"
                f"{availability_sentence + '。' if availability_sentence else ''}希望在实际工作中持续提升业务理解和分析能力。\n\n"
                "简历随信附上，感谢您的阅读，期待有机会进一步沟通！\n\n"
                "此致\n敬礼！\n"
                f"{name}"
            )
            contact_lines = [
                f"姓名：{name}" if name else "",
                f"电话/微信：{self.application_profile.get('phoneWeChat', '')}" if self.application_profile.get("phoneWeChat") else "",
                f"邮箱：{self.application_profile.get('email', '')}" if self.application_profile.get("email") else "",
            ]
            cover_letter = (
                f"主题：{subject}\n"
                "尊敬的招聘负责人：\n"
                f"您好！我是{name or '应聘者'}{identity_text}，了解到贵公司的「{title}」岗位后，希望申请该职位。\n\n"
                "我希望进一步了解该岗位负责的业务方向、核心分析任务和交付标准，"
                "并结合最相关的项目案例说明我的判断方法、实际行动与交付结果。\n\n"
                f"{availability_sentence + '。' if availability_sentence else ''}"
                "希望在实际工作中持续提升业务理解和分析能力。\n\n"
                "简历随信附上，感谢您的阅读，期待有机会进一步沟通！\n\n"
                "此致\n敬礼！\n"
                + "\n".join(line for line in contact_lines if line)
            )
            return {
                "greeting": greeting,
                "email_subject": subject,
                "email_body": email_body,
                "cover_letter": cover_letter,
                "used_evidence_ids": [],
                "requirement_matches": [],
                "recommended_resume": "",
                "resume_reason": "",
                "generation_mode": "deterministic_fallback",
                "runtime_status": "fallback_missing_candidate_evidence",
                "status": "needs_review",
                "format_version": OUTREACH_FORMAT_VERSION,
            }

        evidence_lines = []
        for item in writing_evidence:
            first_person_claim = _text(item.get("first_person_claim")).strip("，。； ")
            if first_person_claim.startswith("我"):
                evidence_lines.append(f"{first_person_claim}。")
                continue
            summary = self._evidence_sentence(item, role_target)
            if summary:
                evidence_lines.append(f"{self._evidence_lead(item)}{summary}。")
        requirements = application_info.get("requirements", [])
        evidence_block = "".join(evidence_lines)
        work_method = self._work_method(focus)
        email_body = (
            "尊敬的招聘负责人：\n"
            f"您好！我是{name or '应聘者'}{identity_text}，了解到贵公司的「{title}」岗位后，希望申请该职位。\n\n"
            f"{evidence_block}这些经历使我能够围绕{focus}梳理事实、形成判断并推进交付。\n\n"
            f"{availability_sentence + '。' if availability_sentence else ''}希望将这些能力应用于实际业务，并持续提升业务理解和分析能力。\n\n"
            "简历随信附上，感谢您的阅读，期待有机会进一步沟通！\n\n"
            "此致\n敬礼！\n"
            f"{name}"
        )
        identity = "".join(
            part for part in (
                self.application_profile.get("school", ""),
                self.application_profile.get("major", ""),
                self.application_profile.get("degreeYear", ""),
            ) if part
        )
        introduction = f"我是{name}" if name else "我希望申请该岗位"
        if identity:
            introduction += f"，目前就读于{identity}"
        availability_sentence = "，".join(filter(None, (availability_text, f"预计可连续实习{duration}" if duration else "")))
        contact_lines = [
            f"姓名：{name}" if name else "",
            f"电话/微信：{self.application_profile.get('phoneWeChat', '')}" if self.application_profile.get("phoneWeChat") else "",
            f"邮箱：{self.application_profile.get('email', '')}" if self.application_profile.get("email") else "",
        ]
        cover_letter = (
            f"主题：{subject}\n"
            "尊敬的招聘负责人：\n"
            f"您好！{introduction}，了解到贵公司的「{title}」岗位后，希望申请该职位。\n\n"
            f"{evidence_block}这些工作要求我从具体目标出发筛选信息、核对关键事实，并把分析结果整理成团队可以继续使用的交付物。"
            "我也会根据反馈及时修正判断，确保结论能够服务于后续决策和执行。\n\n"
            f"{availability_sentence + '。' if availability_sentence else ''}希望将这些能力应用于贵公司的实际业务。若有机会加入，{work_method}，从明确、可验收的任务开始稳定交付。\n\n"
            "简历随信附上，感谢您的阅读，期待有机会进一步沟通！\n\n"
            "此致\n敬礼！\n"
            + "\n".join(line for line in contact_lines if line)
        )
        return {
            "greeting": greeting,
            "email_subject": subject,
            "email_body": email_body,
            "cover_letter": cover_letter,
            "used_evidence_ids": evidence_ids,
            "requirement_matches": [
                f"{_text(item.get('text'))}：对应证据 {writing_evidence[0]['id']}"
                for item in requirements[:2]
            ],
            "recommended_resume": "",
            "resume_reason": "",
            "generation_mode": "deterministic",
            "runtime_status": "not_requested",
            "status": "ready",
            "format_version": OUTREACH_FORMAT_VERSION,
        }


def build_job_card(
    note: dict[str, Any],
    application_info: dict[str, Any],
    *,
    body_present: bool,
) -> dict[str, Any]:
    source_excerpt = next(
        (
            _text(note.get(field))
            for field in ("body", "source_card_text", "card_text_segments", "title", "card_title")
            if _text(note.get(field))
        ),
        "",
    )
    contacts = application_info.get("contacts", [])
    routes = application_info.get("application_routes", [])
    return {
        "title": _text(note.get("title")) or _text(note.get("card_title")) or "未命名岗位",
        "source_url": _text(note.get("note_url"))
        or _text(note.get("search_result_url"))
        or _text(note.get("card_search_result_url")),
        "source_status": _text(note.get("access_status")) or "unknown",
        "parse_basis": "full_body" if body_present else "search_card",
        "source_excerpt": source_excerpt[:280],
        "responsibility_count": len(application_info.get("responsibilities", [])),
        "requirement_count": len(application_info.get("requirements", [])),
        "route_count": len(contacts) + len(routes),
        "status": "generated",
    }


def _record_key(item: dict[str, Any]) -> str:
    return _text(item.get("note_id")) or _text(item.get("note_url")) or _text(item.get("search_result_url")) or _text(item.get("card_search_result_url"))


@dataclass
class PipelineResult:
    payload: dict[str, Any]
    passed: bool


FAILED_DETAIL_STATUSES = {
    "detail_unavailable",
    "detail_timeout",
    "detail_playwright_error",
    "detail_unexpected_error",
    "missing_record",
}


def has_publishable_detail(record: dict[str, Any]) -> bool:
    """Only full detail records may enter the AI/publication pipeline."""
    return bool(_text(record.get("body"))) and _text(record.get("access_status")) not in FAILED_DETAIL_STATUSES


class ApplicationIntelligencePipeline:
    def __init__(
        self,
        profile: dict[str, Any],
        now: datetime | None = None,
        *,
        runtime_agent: Any = None,
        runtime_required: bool = False,
        runtime_initialization_error: str = "",
    ):
        self.profile = profile
        self.now = now or datetime.now(SHANGHAI)
        if self.now.tzinfo is None:
            self.now = self.now.replace(tzinfo=SHANGHAI)
        else:
            self.now = self.now.astimezone(SHANGHAI)
        self.info_agent = ApplicationInfoAgent()
        self.fit_agent = FitEvidenceAgent(profile)
        self.writer_agent = OutreachWriterAgent(profile)
        self.runtime_agent = runtime_agent
        self.runtime_required = runtime_required
        self.runtime_initialization_error = runtime_initialization_error

    def run(self, cards: list[dict[str, Any]], notes: list[dict[str, Any]]) -> PipelineResult:
        note_by_key = {_record_key(note): note for note in notes if _record_key(note)}
        unique_cards: list[dict[str, Any]] = []
        seen_card_keys: set[str] = set()
        for card in cards:
            key = _record_key(card)
            if key and key in seen_card_keys:
                continue
            if key:
                seen_card_keys.add(key)
            unique_cards.append(card)
        merged: list[tuple[dict[str, Any], bool]] = []
        for card in unique_cards:
            key = _record_key(card)
            note = note_by_key.get(key)
            if note is None and _text(card.get("note_id")):
                note = next((candidate for candidate in notes if _text(candidate.get("note_id")) == _text(card.get("note_id"))), None)
            if note is not None and has_publishable_detail(note):
                merged.append(({**card, **note}, True))
        # When cards are present they are the authoritative checkpoint boundary.
        # Resume directories can briefly contain notes copied from an older card
        # set, so appending unmatched notes would create duplicate/stale job cards.
        if not cards:
            merged.extend((note, True) for note in notes if has_publishable_detail(note))

        results: list[dict[str, Any]] = []
        unparsed_times = 0
        empty_bodies = 0
        missing_scraped_at = 0

        for note, has_record in merged:
            collected_at, collection_method = _aware_datetime(note.get("scraped_at"), self.now)
            if collection_method != "scraped_at":
                missing_scraped_at += 1
            raw_time, source_field = _time_source(note)
            normalized_time = normalize_publish_time(raw_time, collected_at, source_field) if raw_time else {
                "raw": "",
                "value": "",
                "timezone": "Asia/Shanghai",
                "precision": "unknown",
                "method": "missing",
                "is_estimated": False,
                "source_field": "",
                "collected_at": collected_at.isoformat(timespec="seconds"),
            }
            if raw_time and not normalized_time["value"]:
                unparsed_times += 1
            info = self.info_agent.run(note)
            fit = self.fit_agent.run(note, info["requirements"])
            outreach = self.writer_agent.run(note, info, fit)

            access_status = _text(note.get("access_status"))
            body_present = has_publishable_detail(note)
            if not body_present:
                empty_bodies += 1
                outreach.update(
                    status="needs_review",
                    runtime_status="fallback_missing_job_body",
                )
            job_card = build_job_card(note, info, body_present=body_present)
            outreach_generated = all(
                _text(outreach.get(field))
                for field in ("greeting", "email_subject", "email_body", "cover_letter")
            )
            results.append(
                {
                    "note_id": _text(note.get("note_id")),
                    "title": _text(note.get("title")) or _text(note.get("card_title")),
                    "note_url": _text(note.get("note_url")) or _text(note.get("search_result_url")) or _text(note.get("card_search_result_url")),
                    "body": _text(note.get("body")),
                    "access_status": access_status or ("missing_record" if not has_record else "unknown"),
                    "collected_at": collected_at.isoformat(timespec="seconds"),
                    "collection_time_source": collection_method,
                    "publish_time": normalized_time,
                    "media": build_media(note),
                    "job_card": job_card,
                    "application_info": info,
                    "fit_evidence": fit,
                    "outreach": outreach,
                    "quality": {
                        "discovered_record_present": has_record,
                        "body_present": body_present,
                        "time_normalized_when_present": not raw_time or bool(normalized_time["value"]),
                        "provenance_valid": True,
                        "job_card_generated": True,
                        "outreach_generated": outreach_generated,
                    },
                }
            )

        if self.runtime_agent is not None:
            runtime_report = self.runtime_agent.enrich(results).as_dict()
        elif self.runtime_required:
            eligible = sum(1 for item in results if item["quality"]["body_present"])
            for record in results:
                if record["quality"]["body_present"]:
                    record["outreach"].update(
                        generation_mode="deterministic_fallback",
                        runtime_status="failed",
                        status="blocked_codex_runtime",
                    )
            runtime_report = {
                "enabled": True,
                "status": "failed",
                "cli": "",
                "prompt_version": PROMPT_VERSION,
                "requested": eligible,
                "generated": 0,
                "cached": 0,
                "failed": eligible,
                "failures": [{"note_id": "runtime", "error": self.runtime_initialization_error}],
            }
        else:
            runtime_report = {
                "enabled": False,
                "status": "disabled",
                "cli": "",
                "prompt_version": PROMPT_VERSION,
                "requested": 0,
                "generated": 0,
                "cached": 0,
                "failed": 0,
                "failures": [],
            }

        invalid_provenance = 0
        blocked_drafts = 0
        job_cards_generated = 0
        outreach_drafts_generated = 0
        valid_evidence_ids = {item["id"] for item in self.fit_agent.evidence}
        for record in results:
            outreach = record["outreach"]
            if record.get("job_card", {}).get("status") == "generated":
                job_cards_generated += 1
            if all(_text(outreach.get(field)) for field in ("greeting", "email_subject", "email_body", "cover_letter")):
                outreach_drafts_generated += 1
            if outreach["status"] != "ready":
                blocked_drafts += 1
            record_valid = not any(
                item not in valid_evidence_ids for item in outreach.get("used_evidence_ids", [])
            )
            for item in (
                record["application_info"]["contacts"]
                + record["application_info"]["application_routes"]
                + record["application_info"]["requirements"]
                + record["application_info"]["responsibilities"]
            ):
                if not item.get("source_field") or not item.get("evidence"):
                    record_valid = False
            record["quality"]["provenance_valid"] = record_valid
            if not record_valid:
                invalid_provenance += 1

        discovered_count = len(unique_cards) if cards else len(notes)
        record_keys = {_record_key(record) for record in results if _record_key(record)}
        card_keys = {_record_key(card) for card in cards if _record_key(card)}
        covered_discovered = len(card_keys & record_keys) if cards else len(record_keys)
        missing_records = max(0, discovered_count - covered_discovered)
        issues = []
        checks = {
            "all_discovered_notes_have_records": missing_records == 0 and covered_discovered == discovered_count,
            "all_records_have_bodies": empty_bodies == 0,
            "all_relative_or_absolute_times_normalized": unparsed_times == 0,
            "all_records_have_source_collection_time": missing_scraped_at == 0,
            "all_extractions_and_drafts_have_provenance": invalid_provenance == 0,
            "candidate_evidence_loaded": bool(self.fit_agent.evidence),
            "all_scraped_jobs_have_job_cards": job_cards_generated == len(results),
            "all_scraped_jobs_have_application_copy": outreach_drafts_generated == len(results),
            "all_outreach_drafts_ready": blocked_drafts == 0,
        }
        if self.runtime_required:
            checks["all_outreach_generated_by_codex_runtime"] = (
                runtime_report["failed"] == 0
                and runtime_report["requested"] == runtime_report["generated"] + runtime_report["cached"]
            )
        messages = {
            "all_discovered_notes_have_records": (
                f"{missing_records} discovered cards are waiting for a full detail body before publication"
            ),
            "all_records_have_bodies": f"{empty_bodies} discovered notes have no full body",
            "all_relative_or_absolute_times_normalized": f"{unparsed_times} supplied time labels could not be normalized",
            "all_records_have_source_collection_time": f"{missing_scraped_at} records lack source scraped_at and use a marked fallback",
            "all_extractions_and_drafts_have_provenance": f"{invalid_provenance} provenance references are invalid",
            "candidate_evidence_loaded": "candidate profile contains no usable resume or GitHub evidence",
            "all_scraped_jobs_have_job_cards": f"{len(results) - job_cards_generated} scraped jobs have no generated job card",
            "all_scraped_jobs_have_application_copy": f"{len(results) - outreach_drafts_generated} scraped jobs have no editable application copy",
            "all_outreach_drafts_ready": f"{blocked_drafts} outreach drafts are blocked",
            "all_outreach_generated_by_codex_runtime": (
                f"Codex Runtime generated or reused {runtime_report['generated'] + runtime_report['cached']} "
                f"of {runtime_report['requested']} eligible drafts"
            ),
        }
        for check, passed in checks.items():
            if not passed:
                issues.append({"check": check, "message": messages[check]})
        passed = all(checks.values())
        summary = {
            "passed": passed,
            "generated_at": self.now.isoformat(timespec="seconds"),
            "timezone": "Asia/Shanghai",
            "discovered_count": discovered_count,
            "record_count": len(results),
            "covered_discovered_count": covered_discovered,
            "body_count": sum(1 for item in results if item["quality"]["body_present"]),
            "job_cards_generated": job_cards_generated,
            "application_copy_generated": outreach_drafts_generated,
            "generation_coverage_rate": (outreach_drafts_generated / len(results)) if results else 1.0,
            "coverage_rate": (covered_discovered / discovered_count) if discovered_count else 1.0,
            "body_coverage_rate": (sum(1 for item in results if item["quality"]["body_present"]) / discovered_count) if discovered_count else 1.0,
            "checks": checks,
            "issues": issues,
        }
        return PipelineResult(
            payload={
                "schema_version": "1.4",
                "candidate_name": _candidate_name(self.profile),
                "publication_contract": {
                    "mode": "card_body_atomic",
                    "candidate_count": discovered_count,
                    "published_count": len(results),
                    "pending_body_count": missing_records,
                    "ai_runs_after_body_collection": True,
                },
                "agents": [
                    {"id": "coverage-agent", "status": "completed", "output": "discovery and body coverage"},
                    {"id": "time-agent", "status": "completed", "output": "Asia/Shanghai publication time"},
                    {"id": "application-info-agent", "status": "completed", "output": "responsibilities, requirements, contacts, and routes with provenance"},
                    {"id": "fit-evidence-agent", "status": "completed", "output": "resume and GitHub evidence matches"},
                    {"id": "outreach-writer-agent", "status": runtime_report["status"], "output": "per-note greeting, email, and cover letter through Codex Runtime"},
                    {"id": "quality-gate-agent", "status": "passed" if passed else "failed", "output": "coverage and provenance checks"},
                ],
                "quality_gate": summary,
                "codex_runtime": runtime_report,
                "records": results,
            },
            passed=passed,
        )


def _load_json(path: Path, default: Any) -> Any:
    if not path.exists():
        return default
    return json.loads(path.read_text(encoding="utf-8"))


def find_notes_path(output_dir: Path) -> Path:
    for name in ("xiaohongshu_notes_latest.json", "xiaohongshu_notes_latest_dedup.json"):
        candidate = output_dir / name
        if candidate.exists():
            return candidate
    raise FileNotFoundError(f"No latest notes JSON found under {output_dir}")


def run_pipeline(
    output_dir: Path,
    candidate_profile_path: Path,
    now: datetime | None = None,
    *,
    use_codex_runtime: bool = False,
    codex_cli_bin: str = "",
    codex_batch_size: int = 8,
    codex_timeout_seconds: int = 300,
    persist: bool = True,
) -> PipelineResult:
    cards = _load_json(output_dir / "xiaohongshu_cards_latest.json", [])
    notes = _load_json(find_notes_path(output_dir), [])
    profile = _load_json(candidate_profile_path, {})
    if not isinstance(cards, list) or not isinstance(notes, list) or not isinstance(profile, dict):
        raise ValueError("Cards and notes must be JSON arrays; candidate profile must be a JSON object")
    runtime_agent = None
    runtime_error = ""
    if use_codex_runtime:
        try:
            from codex_runtime_outreach import CodexRuntimeOutreachAgent

            runtime_agent = CodexRuntimeOutreachAgent(
                output_dir,
                candidate_name=_candidate_name(profile),
                candidate_profile=_candidate_application_profile(profile),
                cli_bin=codex_cli_bin,
                batch_size=codex_batch_size,
                timeout_seconds=codex_timeout_seconds,
            )
        except (OSError, ValueError) as error:
            runtime_error = str(error)
    result = ApplicationIntelligencePipeline(
        profile,
        now=now,
        runtime_agent=runtime_agent,
        runtime_required=use_codex_runtime,
        runtime_initialization_error=runtime_error,
    ).run(cards, notes)
    if persist:
        write_pipeline_artifacts(output_dir, result.payload)
    return result


def write_pipeline_artifacts(output_dir: Path, payload: dict[str, Any]) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)
    json_path = output_dir / "application_intelligence.json"
    atomic_write_json(json_path, payload)

    summary = payload["quality_gate"]
    summary_path = output_dir / "application_intelligence_summary.json"
    atomic_write_json(summary_path, summary)
    coverage_path = output_dir / "coverage_report.json"
    atomic_write_json(coverage_path, summary)

    csv_path = output_dir / "application_intelligence.csv"
    with csv_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=(
                "note_id",
                "title",
                "note_url",
                "collected_at",
                "publish_time_raw",
                "publish_time_normalized",
                "publish_time_precision",
                "publish_time_is_estimated",
                "contact_values",
                "application_routes",
                "responsibilities",
                "requirements",
                "fit_evidence_ids",
                "fit_match_confidence",
                "greeting",
                "email_subject",
                "email_body",
                "cover_letter",
                "recommended_resume",
                "generation_mode",
                "runtime_status",
                "quality_status",
            ),
        )
        writer.writeheader()
        for record in payload["records"]:
            contacts = record["application_info"]["contacts"]
            routes = record["application_info"]["application_routes"]
            responsibilities = record["application_info"]["responsibilities"]
            requirements = record["application_info"]["requirements"]
            writer.writerow(
                {key: _sheet_safe(value) for key, value in {
                    "note_id": record["note_id"],
                    "title": record["title"],
                    "note_url": record["note_url"],
                    "collected_at": record["collected_at"],
                    "publish_time_raw": record["publish_time"]["raw"],
                    "publish_time_normalized": record["publish_time"]["value"],
                    "publish_time_precision": record["publish_time"]["precision"],
                    "publish_time_is_estimated": record["publish_time"]["is_estimated"],
                    "contact_values": " | ".join(f"{item['type']}:{item['value']}" for item in contacts),
                    "application_routes": " | ".join(item["type"] for item in routes),
                    "responsibilities": " | ".join(item["text"] for item in responsibilities),
                    "requirements": " | ".join(item["text"] for item in requirements),
                    "fit_evidence_ids": " | ".join(item["id"] for item in record["fit_evidence"]),
                    "fit_match_confidence": " | ".join(str(item.get("match_confidence", "")) for item in record["fit_evidence"]),
                    "greeting": record["outreach"]["greeting"],
                    "email_subject": record["outreach"]["email_subject"],
                    "email_body": record["outreach"]["email_body"],
                    "cover_letter": record["outreach"].get("cover_letter", ""),
                    "recommended_resume": record["outreach"].get("recommended_resume", ""),
                    "generation_mode": record["outreach"].get("generation_mode", ""),
                    "runtime_status": record["outreach"].get("runtime_status", ""),
                    "quality_status": "pass" if all(record["quality"].values()) else "needs_review",
                }.items()}
            )

    markdown_path = output_dir / "application_intelligence_report.md"
    if payload.get("analysis_mode") == "general":
        insights = payload.get("content_insights") if isinstance(payload.get("content_insights"), dict) else {}
        research = payload.get("content_research") if isinstance(payload.get("content_research"), dict) else {}
        lines = [
            f"# {payload.get('keyword') or '关键词'}跨样本内容洞察报告",
            "",
            f"- 研究场景：{research.get('label') or research.get('preset') or '非岗位内容研究'}",
            f"- 样本总数：{int(insights.get('sampleSize') or 0)}",
            f"- 原始图文就绪：{int(insights.get('sourceReady') or 0)}",
            f"- 证据门禁通过：{int(insights.get('groundedRecords') or 0)}",
            f"- 可验证分析覆盖率：{float(insights.get('coverageRate') or 0):.1f}%",
            "",
            f"> {insights.get('methodNote') or '结论仅来自可回溯的原始图文证据。'}",
            "",
            "## 高频主题",
        ]
        topics = insights.get("topTopics") if isinstance(insights.get("topTopics"), list) else []
        if topics:
            for topic in topics:
                lines.append(f"- {topic.get('label') or '未命名主题'}：{int(topic.get('count') or 0)} 条，占证据通过样本 {float(topic.get('share') or 0):.1f}%")
                for evidence in topic.get("evidence", [])[:3]:
                    lines.append(f"  - 《{evidence.get('title') or '未命名内容'}》：{evidence.get('quote') or ''}")
        else:
            lines.append("- 当前没有通过原文证据门禁的高频主题。")
        for module in insights.get("modules", []):
            lines.extend([
                "",
                f"## {module.get('title') or '研究问题'}",
                "",
                f"研究问题：{module.get('question') or '未设置'}",
                f"证据覆盖：{int(module.get('recordCount') or 0)} 条（{float(module.get('coverageRate') or 0):.1f}%）",
                "",
            ])
            findings = module.get("findings") if isinstance(module.get("findings"), list) else []
            if not findings:
                lines.append("- 当前样本没有形成可复核结论。")
                continue
            for finding in findings:
                lines.append(f"- {finding.get('label') or '未命名发现'}：{int(finding.get('count') or 0)} 条")
                for evidence in finding.get("evidence", [])[:3]:
                    lines.append(f"  - 《{evidence.get('title') or '未命名内容'}》：{evidence.get('quote') or ''}")
    else:
        lines = [
            "# Application Intelligence Quality Report",
            "",
            f"- Gate: {'PASS' if summary['passed'] else 'FAIL'}",
            f"- Discovered: {summary['discovered_count']}",
            f"- Extracted records: {summary['record_count']}",
            f"- Full bodies: {summary['body_count']}",
            f"- Discovery coverage: {summary['coverage_rate']:.1%}",
            f"- Body coverage: {summary['body_coverage_rate']:.1%}",
            "",
            "## Checks",
        ]
        lines.extend(f"- [{'x' if passed else ' '}] {name}" for name, passed in summary["checks"].items())
        if summary["issues"]:
            lines.extend(["", "## Issues"])
            lines.extend(f"- {item['message']}" for item in summary["issues"])
    markdown_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    _write_xlsx(output_dir / "application_intelligence.xlsx", payload)


def _write_xlsx(path: Path, payload: dict[str, Any]) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to write application_intelligence.xlsx") from exc

    workbook = Workbook()
    applications = workbook.active
    applications.title = "Applications"
    application_headers = [
        "note_id",
        "title",
        "note_url",
        "access_status",
        "collected_at",
        "publish_time_raw",
        "publish_time_normalized",
        "publish_time_precision",
        "publish_time_is_estimated",
        "body",
        "greeting",
        "email_subject",
        "email_body",
        "cover_letter",
        "recommended_resume",
        "resume_reason",
        "generation_mode",
        "runtime_status",
        "draft_status",
    ]
    applications.append(application_headers)
    for record in payload["records"]:
        applications.append(
            [_sheet_safe(value) for value in [
                record["note_id"],
                record["title"],
                record["note_url"],
                record["access_status"],
                record["collected_at"],
                record["publish_time"]["raw"],
                record["publish_time"]["value"],
                record["publish_time"]["precision"],
                record["publish_time"]["is_estimated"],
                record["body"],
                record["outreach"]["greeting"],
                record["outreach"]["email_subject"],
                record["outreach"]["email_body"],
                record["outreach"].get("cover_letter", ""),
                record["outreach"].get("recommended_resume", ""),
                record["outreach"].get("resume_reason", ""),
                record["outreach"].get("generation_mode", ""),
                record["outreach"].get("runtime_status", ""),
                record["outreach"]["status"],
            ]]
        )

    contacts = workbook.create_sheet("Contacts")
    contacts.append(["note_id", "type", "value", "source_field", "evidence", "offset_start", "offset_end"])
    requirements = workbook.create_sheet("Requirements")
    requirements.append(["note_id", "requirement", "source_field", "evidence", "offset_start", "offset_end", "classification_confidence", "classification_basis"])
    responsibilities = workbook.create_sheet("Responsibilities")
    responsibilities.append(["note_id", "responsibility", "source_field", "evidence", "offset_start", "offset_end", "classification_confidence", "classification_basis"])
    evidence = workbook.create_sheet("Fit Evidence")
    evidence.append(["note_id", "evidence_id", "category", "label", "detail", "source", "match_score", "match_confidence", "matched_terms", "match_basis"])
    for record in payload["records"]:
        for item in record["application_info"]["contacts"] + record["application_info"]["application_routes"]:
            contacts.append(
                [_sheet_safe(value) for value in [
                    record["note_id"],
                    item["type"],
                    item["value"],
                    item["source_field"],
                    item["evidence"],
                    item.get("offset_start", -1),
                    item.get("offset_end", -1),
                ]]
            )
        for item in record["application_info"]["requirements"]:
            requirements.append(
                [_sheet_safe(value) for value in [
                    record["note_id"],
                    item["text"],
                    item["source_field"],
                    item["evidence"],
                    item.get("offset_start", -1),
                    item.get("offset_end", -1),
                    item.get("classification_confidence", 0),
                    " | ".join(item.get("classification_basis", [])),
                ]]
            )
        for item in record["application_info"]["responsibilities"]:
            responsibilities.append(
                [_sheet_safe(value) for value in [
                    record["note_id"],
                    item["text"],
                    item["source_field"],
                    item["evidence"],
                    item.get("offset_start", -1),
                    item.get("offset_end", -1),
                    item.get("classification_confidence", 0),
                    " | ".join(item.get("classification_basis", [])),
                ]]
            )
        for item in record["fit_evidence"]:
            evidence.append(
                [_sheet_safe(value) for value in [
                    record["note_id"],
                    item["id"],
                    item["category"],
                    item["label"],
                    item["detail"],
                    item["source"],
                    item["match_score"],
                    item.get("match_confidence", 0),
                    " | ".join(item.get("matched_terms", [])),
                    item.get("match_basis", ""),
                ]]
            )

    quality = workbook.create_sheet("Quality Gate")
    quality.append(["check", "passed", "detail"])
    issue_by_check = {
        item.get("check") or item.get("code") or "unknown": item.get("message", "")
        for item in payload["quality_gate"]["issues"]
    }
    for check, passed in payload["quality_gate"]["checks"].items():
        quality.append([check, passed, issue_by_check.get(check, "")])
    quality.append([])
    for field in ("generated_at", "timezone", "discovered_count", "record_count", "covered_discovered_count", "body_count", "coverage_rate", "body_coverage_rate"):
        quality.append([field, payload["quality_gate"][field], ""])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for column in sheet.columns:
            letter = column[0].column_letter
            max_width = max((len(str(cell.value or "")) for cell in column), default=8)
            sheet.column_dimensions[letter].width = min(max(max_width + 2, 10), 48)

    temporary_path = path.with_suffix(".xlsx.tmp")
    workbook.save(temporary_path)
    temporary_path.replace(path)
from artifact_io import atomic_write_json
